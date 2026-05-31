from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_socketio import SocketIO, emit, join_room
from flask_mail import Mail, Message
from apscheduler.schedulers.background import BackgroundScheduler
from flask_sqlalchemy import SQLAlchemy
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
import json
from datetime import datetime, timedelta
from urllib.parse import urljoin
from urllib.parse import urlparse
import socket
import ipaddress
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import io
import hashlib
from flask import send_file
import secrets
import re
import threading
from sqlalchemy import text
from flask_caching import Cache
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# ===================== NEW LIBRARIES INSTALLED =====================
import requests
from bs4 import BeautifulSoup
try:
    import whois
except ImportError:
    whois = None

try:
    import dns.resolver
except ImportError:
    dns = None

try:
    import bleach
except ImportError:
    bleach = None

try:
    import phonenumbers
except ImportError:
    phonenumbers = None

try:
    import magic
except ImportError:
    magic = None

try:
    import numpy as np
except ImportError:
    np = None

try:
    from web3 import Web3
except Exception:
    Web3 = None

# OR if that fails, use:
# import magic as filemagic
from email_validator import validate_email, EmailNotValidError
import logging

from chatbot_engine import ChatbotEngine, get_ai_response

# Scikit-learn / pandas for recommendation engine
SKLEARN_AVAILABLE = False
PANDAS_AVAILABLE = False
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    print("Pandas is not installed; recommendation engine will use fallback logic.")

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    SKLEARN_AVAILABLE = True
except ImportError:
    print("Scikit-learn is not installed; recommendation engine will fall back to keyword matching.")

# ML libraries for job content analysis
try:
    from sentence_transformers import SentenceTransformer
    ML_AVAILABLE = True
except ImportError:
    print("ML libraries not available. Install with: pip install sentence-transformers")
    ML_AVAILABLE = False

# Transformer pipeline (optional)
try:
    from transformers import pipeline
    TRANSFORMERS_AVAILABLE = True
except Exception:
    TRANSFORMERS_AVAILABLE = False

def safe_print(*values, **kwargs):
    sanitized = []
    for value in values:
        if isinstance(value, str):
            sanitized.append(value.encode('ascii', 'ignore').decode('ascii'))
        else:
            sanitized.append(str(value))
    print(*sanitized, **kwargs)


# Initialize ML models
job_similarity_model = None
sentence_encoder = None
if ML_AVAILABLE:
    try:
        sentence_encoder = SentenceTransformer('all-MiniLM-L6-v2')
        safe_print("Sentence transformer loaded")
    except Exception as e:
        safe_print(f"Sentence transformer loading failed: {e}")
        sentence_encoder = None
        ML_AVAILABLE = False

# Initialize ML models
job_classifier = None
sentence_encoder = None

if ML_AVAILABLE and TRANSFORMERS_AVAILABLE:
    try:
        job_classifier = pipeline("text-classification",
                                  model="facebook/bart-large-mnli",
                                  return_all_scores=True)
        sentence_encoder = SentenceTransformer('all-MiniLM-L6-v2')
        safe_print("ML models loaded successfully")
    except Exception as e:
        safe_print(f"ML model loading failed: {e}")
        ML_AVAILABLE = False
elif ML_AVAILABLE and not TRANSFORMERS_AVAILABLE:
    safe_print("transformers.pipeline not available; install transformers to enable job classification")

app = Flask(__name__)
app.secret_key = "portal_secret_key_2026_change_this_in_production"

# Neon PostgreSQL connection
DATABASE_URL = "postgresql://neondb_owner:npg_71uZUrhQmBJa@ep-hidden-morning-apcsf9y6-pooler.c-7.us-east-1.aws.neon.tech/neondb?sslmode=require"

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

os.environ['DATABASE_URL'] = DATABASE_URL

db = SQLAlchemy(app)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# SQLAlchemy (optional models integration)
from models import init_models

# ===================== ENHANCED CONFIGURATIONS =====================

# Cache setup (for faster performance)
cache = Cache(app, config={'CACHE_TYPE': 'simple'})

# Rate limiting (prevent spam)
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Content moderation settings
BANNED_WORDS = ['scam', 'fraud', 'hate', 'violence', 'kill', 'abuse', 'harassment', 'racist', 'stupid', 'idiot']
SUSPICIOUS_PATTERNS = [
    r'\b(credit card|ssn|password|bank account|bitcoin|paypal)\b',
    r'\b(buy|sell|cheap|discount)\b.*\b(weed|drugs|pills)\b'
]

# Company verification thresholds
VERIFICATION_THRESHOLDS = {
    'auto_approve': 70,  # Score >= 70 = auto-verified
    'manual_review': 40,  # Score 40-69 = needs admin review
    'reject': 39  # Score <= 39 = rejected
}

# ===================== EMAIL CONFIG =====================
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'tbosetumi@gmail.com'    # REPLACE
app.config['MAIL_PASSWORD'] = 'katali5931'       # REPLACE
mail = Mail(app)

# ===================== AFRICA'S TALKING SMS CONFIG =====================
# Replace with your Africa's Talking credentials
AT_USERNAME = 'sandbox'  # Your Africa's Talking username
# ===================== SOCKETIO =====================
socketio = SocketIO(app, cors_allowed_origins="*")

# ===================== UPLOAD FOLDERS =====================
UPLOAD_FOLDER = 'static/uploads/cv'
PROFILE_PIC_FOLDER = 'static/uploads/profiles'
STATUS_MEDIA_FOLDER = 'static/uploads/statuses'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROFILE_PIC_FOLDER, exist_ok=True)
os.makedirs(STATUS_MEDIA_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROFILE_PIC_FOLDER'] = PROFILE_PIC_FOLDER
app.config['STATUS_MEDIA_FOLDER'] = STATUS_MEDIA_FOLDER

scheduler = BackgroundScheduler()

def get_db():
    conn = sqlite3.connect("database.db", timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA busy_timeout=30000")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def safe_print(*values, **kwargs):
    sanitized = []
    for value in values:
        if isinstance(value, str):
            sanitized.append(value.encode('ascii', 'ignore').decode('ascii'))
        else:
            sanitized.append(str(value))
    print(*sanitized, **kwargs)


def cleanup_duplicate_student_numbers(conn):
    duplicate_rows = conn.execute("""
        SELECT student_number, GROUP_CONCAT(id) as ids
        FROM users
        WHERE student_number IS NOT NULL
        GROUP BY student_number
        HAVING COUNT(*) > 1
    """).fetchall()

    cleared = 0
    for duplicate in duplicate_rows:
        user_ids = [int(user_id) for user_id in duplicate['ids'].split(',')]
        keep_user_id = min(user_ids)
        duplicate_user_ids = [user_id for user_id in user_ids if user_id != keep_user_id]
        if duplicate_user_ids:
            placeholders = ','.join('?' for _ in duplicate_user_ids)
            conn.execute(
                f"UPDATE users SET student_number = NULL WHERE id IN ({placeholders})",
                duplicate_user_ids
            )
            logger.warning("Cleared duplicate student_number %s from user ids %s", duplicate['student_number'], duplicate_user_ids)
            cleared += len(duplicate_user_ids)
    return cleared


def repair_corrupt_text_files(root_dir):
    repaired = []
    for directory, _, filenames in os.walk(root_dir):
        for filename in filenames:
            if not filename.endswith(('.html', '.js', '.css', '.json', '.txt')):
                continue
            file_path = os.path.join(directory, filename)
            try:
                raw_bytes = open(file_path, 'rb').read()
            except OSError:
                continue
            try:
                raw_bytes.decode('utf-8')
            except UnicodeDecodeError:
                repaired_encoding = None
                for fallback_encoding in ('cp1252', 'latin-1'):
                    try:
                        repaired_encoding = raw_bytes.decode(fallback_encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                if repaired_encoding is None:
                    repaired_encoding = raw_bytes.decode('latin-1', errors='replace')
                with open(file_path, 'w', encoding='utf-8') as repaired_file:
                    repaired_file.write(repaired_encoding)
                repaired.append(file_path)
    return repaired


def normalize_student_number(student_number):
    normalized = str(student_number or "").strip()
    return re.sub(r"\D", "", normalized)


def is_valid_student_number(student_number):
    normalized = normalize_student_number(student_number)
    return bool(re.fullmatch(r"\d{9}", normalized))


def ensure_user_student_number_schema(conn):
    existing_user_cols = [row[1] for row in conn.execute("PRAGMA table_info(users)").fetchall()]
    if 'student_number' not in existing_user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN student_number TEXT")
    conn.execute("UPDATE users SET student_number = student_id WHERE student_number IS NULL AND student_id IS NOT NULL")
    duplicate_student_numbers = conn.execute("""
        SELECT student_number, COUNT(*) as cnt
        FROM users
        WHERE student_number IS NOT NULL
        GROUP BY student_number
        HAVING cnt > 1
    """).fetchall()
    for duplicate in duplicate_student_numbers:
        logger.warning("Duplicate student_number detected during migration: %s (%s rows)", duplicate['student_number'], duplicate['cnt'])
    try:
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_student_number_unique ON users(student_number)")
    except sqlite3.IntegrityError as exc:
        logger.warning("Unable to create unique student_number index: %s", exc)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png'}

def send_email(recipient, subject, body, html=None):
    try:
        msg = Message(subject, recipients=[recipient])
        msg.body = body
        if html:
            msg.html = html
        mail.send(msg)
        print(f"Email sent to {recipient}")
    except Exception as e:
        print(f"Email error: {e}")

# Repair any corrupted text files during startup so Jinja and static assets decode cleanly.
repaired_files = repair_corrupt_text_files(os.path.join(BASE_DIR, 'templates'))
if repaired_files:
    logger.warning("Repaired text encodings in %s files: %s", len(repaired_files), repaired_files)

# ====================== INITIALIZE DATABASE ======================
def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        email TEXT UNIQUE NOT NULL,
        phone TEXT UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL,
        student_id TEXT,
        student_number TEXT,
        profile_picture TEXT DEFAULT 'default-avatar.png',
        reset_token TEXT,
        reset_token_expiry TIMESTAMP,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        academic_data_pending INTEGER DEFAULT 0,
        academic_data_last_attempt TIMESTAMP,
        academic_data_retry_count INTEGER DEFAULT 0,
        academic_data_raw TEXT
    )""")
    existing_user_cols = [row[1] for row in conn.execute("PRAGMA table_info(users)").fetchall()]
    if 'phone' not in existing_user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN phone TEXT")
    ensure_user_student_number_schema(conn)
    cleanup_duplicate_student_numbers(conn)
    for col, col_type in [
        ('academic_data_pending', 'INTEGER DEFAULT 0'),
        ('academic_data_last_attempt', 'TIMESTAMP'),
        ('academic_data_retry_count', 'INTEGER DEFAULT 0'),
        ('academic_data_raw', 'TEXT'),
        ('is_alumni', 'INTEGER DEFAULT 0'),
        ('expected_graduation', 'INTEGER')
    ]:
        if col not in existing_user_cols:
            conn.execute(f"ALTER TABLE users ADD COLUMN {col} {col_type}")
    cur.execute("""CREATE TABLE IF NOT EXISTS job_interests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_id INTEGER NOT NULL,
        student_id INTEGER NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        status TEXT DEFAULT 'pending',
        feedback TEXT,
        UNIQUE(job_id, student_id)
    )""")
    existing_cols = [row[1] for row in conn.execute("PRAGMA table_info(job_interests)").fetchall()]
    if 'updated_at' not in existing_cols:
        conn.execute("ALTER TABLE job_interests ADD COLUMN updated_at TIMESTAMP")
        conn.execute("UPDATE job_interests SET updated_at = ? WHERE updated_at IS NULL", (datetime.now(),))
    cur.execute("""CREATE TABLE IF NOT EXISTS mentorships (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        alumni_id INTEGER NOT NULL,
        title TEXT NOT NULL,
        description TEXT,
        duration TEXT,
        skills_needed TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS sis_students (
        student_id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        course TEXT NOT NULL,
        faculty TEXT NOT NULL,
        gpa REAL,
        graduation_year INTEGER
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER UNIQUE NOT NULL,
        course TEXT,
        faculty TEXT,
        gpa REAL,
        graduation_year INTEGER,
        internship_status TEXT DEFAULT 'Seeking Internship',
        skills TEXT,
        results TEXT,
        cv_filename TEXT,
        certificate_hash TEXT,
        certificate_verified INTEGER DEFAULT 0
    )""")
    existing_student_cols = [row[1] for row in conn.execute("PRAGMA table_info(students)").fetchall()]
    if 'certificate_hash' not in existing_student_cols:
        conn.execute("ALTER TABLE students ADD COLUMN certificate_hash TEXT")
    if 'certificate_verified' not in existing_student_cols:
        conn.execute("ALTER TABLE students ADD COLUMN certificate_verified INTEGER DEFAULT 0")
    cur.execute("""CREATE TABLE IF NOT EXISTS recruiters (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER UNIQUE NOT NULL,
        company_name TEXT,
        industry TEXT,
        location TEXT,
        website TEXT,
        website_title TEXT,
        website_description TEXT,
        website_logo_url TEXT,
        social_links TEXT,
        verified INTEGER DEFAULT 0,
        verification_score INTEGER DEFAULT 0,
        verification_status TEXT DEFAULT 'pending'
    )""")
    existing_recruiter_cols = [row[1] for row in conn.execute("PRAGMA table_info(recruiters)").fetchall()]
    if 'website_title' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN website_title TEXT")
    if 'website_description' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN website_description TEXT")
    if 'website_logo_url' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN website_logo_url TEXT")
    if 'description' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN description TEXT")
    if 'industry' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN industry TEXT")
    if 'logo_url' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN logo_url TEXT")
    if 'social_links' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN social_links TEXT")
    if 'company_website' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN company_website TEXT")
    if 'company_email' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN company_email TEXT")
    if 'company_phone' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN company_phone TEXT")
    if 'company_location' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN company_location TEXT")
    if 'last_scrape_at' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN last_scrape_at TIMESTAMP")
    if 'jobs_found_count' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN jobs_found_count INTEGER DEFAULT 0")
    if 'scrape_status' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN scrape_status TEXT DEFAULT 'idle'")
    if 'scrape_message' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN scrape_message TEXT")
    if 'scrape_error' not in existing_recruiter_cols:
        conn.execute("ALTER TABLE recruiters ADD COLUMN scrape_error TEXT")
    cur.execute("""CREATE TABLE IF NOT EXISTS jobs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        recruiter_id INTEGER NOT NULL,
        title TEXT NOT NULL,
        description TEXT,
        requirements TEXT,
        skills_required TEXT,
        location TEXT,
        salary TEXT,
        status TEXT DEFAULT 'pending',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    existing_job_cols = [row[1] for row in conn.execute("PRAGMA table_info(jobs)").fetchall()]
    if 'requirements' not in existing_job_cols:
        conn.execute("ALTER TABLE jobs ADD COLUMN requirements TEXT")
    if 'skills_required' not in existing_job_cols:
        conn.execute("ALTER TABLE jobs ADD COLUMN skills_required TEXT")
    if 'is_active' not in existing_job_cols:
        conn.execute("ALTER TABLE jobs ADD COLUMN is_active INTEGER DEFAULT 1")
    if 'is_filled' not in existing_job_cols:
        conn.execute("ALTER TABLE jobs ADD COLUMN is_filled INTEGER DEFAULT 0")
    if 'updated_at' not in existing_job_cols:
        conn.execute("ALTER TABLE jobs ADD COLUMN updated_at TIMESTAMP")
        # Migration: Set updated_at to current time for existing rows
        conn.execute("UPDATE jobs SET updated_at = ? WHERE updated_at IS NULL", (datetime.now(),))

    cur.execute("""CREATE TABLE IF NOT EXISTS job_recommendations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_id INTEGER NOT NULL,
        student_id INTEGER NOT NULL,
        score REAL NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(job_id, student_id)
    )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_job_recommendations_job ON job_recommendations(job_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_job_recommendations_student ON job_recommendations(student_id)")

    cur.execute("""CREATE TABLE IF NOT EXISTS messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sender_id INTEGER NOT NULL,
        receiver_id INTEGER NOT NULL,
        message TEXT NOT NULL,
        read_status INTEGER DEFAULT 0,
        read_at TIMESTAMP,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    
    cur.execute("""CREATE TABLE IF NOT EXISTS chatbot_conversations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        user_message TEXT NOT NULL,
        bot_response TEXT NOT NULL,
        escalated_to_admin INTEGER DEFAULT 0,
        intent TEXT,
        confidence REAL DEFAULT 0.0,
        session_id TEXT,
        metadata TEXT,
        toxicity_flag INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")

    chatbot_conversation_cols = [row[1] for row in conn.execute("PRAGMA table_info(chatbot_conversations)").fetchall()]
    if 'intent' not in chatbot_conversation_cols:
        conn.execute("ALTER TABLE chatbot_conversations ADD COLUMN intent TEXT")
    if 'confidence' not in chatbot_conversation_cols:
        conn.execute("ALTER TABLE chatbot_conversations ADD COLUMN confidence REAL DEFAULT 0.0")
    if 'session_id' not in chatbot_conversation_cols:
        conn.execute("ALTER TABLE chatbot_conversations ADD COLUMN session_id TEXT")
    if 'metadata' not in chatbot_conversation_cols:
        conn.execute("ALTER TABLE chatbot_conversations ADD COLUMN metadata TEXT")
    if 'toxicity_flag' not in chatbot_conversation_cols:
        conn.execute("ALTER TABLE chatbot_conversations ADD COLUMN toxicity_flag INTEGER DEFAULT 0")

    cur.execute("""CREATE TABLE IF NOT EXISTS chatbot_memory (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id TEXT NOT NULL,
        user_id INTEGER,
        role TEXT NOT NULL,
        content TEXT NOT NULL,
        metadata TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_chatbot_memory_session ON chatbot_memory(session_id, created_at)")

    cur.execute("""CREATE TABLE IF NOT EXISTS chatbot_feedback (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id TEXT,
        user_id INTEGER,
        rating INTEGER,
        comment TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS chatbot_intents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intent TEXT NOT NULL,
        example TEXT,
        confidence REAL DEFAULT 0.0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS chatbot_knowledge (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        question TEXT NOT NULL,
        answer TEXT NOT NULL,
        category TEXT NOT NULL,
        source TEXT DEFAULT 'seeded',
        confidence REAL DEFAULT 0.9,
        tags TEXT,
        embedding TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS broadcast_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        admin_id INTEGER NOT NULL,
        message TEXT NOT NULL,
        sent_to_all INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    
    cur.execute("""CREATE TABLE IF NOT EXISTS projects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        title TEXT NOT NULL,
        description TEXT,
        link TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    
    # Create moderation flags table
    cur.execute("""CREATE TABLE IF NOT EXISTS moderation_flags (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        message TEXT,
        sender_id INTEGER,
        receiver_id INTEGER,
        reason TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        reviewed INTEGER DEFAULT 0
    )""")
    
    # Create message requests table
    cur.execute("""CREATE TABLE IF NOT EXISTS message_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sender_id INTEGER NOT NULL,
        receiver_id INTEGER NOT NULL,
        message TEXT,
        status TEXT DEFAULT 'pending', -- 'pending', 'accepted', 'rejected'
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        responded_at TIMESTAMP
    )""")
    
    # Create scraped jobs table
    cur.execute("""CREATE TABLE IF NOT EXISTS scraped_jobs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source_url TEXT NOT NULL,
        title TEXT NOT NULL,
        description TEXT,
        company TEXT,
        location TEXT,
        salary TEXT,
        job_type TEXT,
        posted_date TIMESTAMP,
        scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        status TEXT DEFAULT 'active'
    )""")
    
    # Create results verification table
    cur.execute("""CREATE TABLE IF NOT EXISTS results_verifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL,
        result_type TEXT,
        results_text TEXT,
        results_file TEXT,
        certificate_hash TEXT,
        status TEXT DEFAULT 'pending', -- 'pending', 'verified', 'rejected'
        admin_notes TEXT,
        verified_by INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        verified_at TIMESTAMP
    )""")
    # Ensure anchor_tx column exists to store blockchain tx hash (optional)
    existing_results_cols = [row[1] for row in conn.execute("PRAGMA table_info(results_verifications)").fetchall()]
    if 'anchor_tx' not in existing_results_cols:
        try:
            conn.execute("ALTER TABLE results_verifications ADD COLUMN anchor_tx TEXT")
        except Exception:
            pass
    existing_results_cols = [row[1] for row in conn.execute("PRAGMA table_info(results_verifications)").fetchall()]
    if 'result_type' not in existing_results_cols:
        conn.execute("ALTER TABLE results_verifications ADD COLUMN result_type TEXT")
    
    # Create session management table for 4-hour persistence
    cur.execute("""CREATE TABLE IF NOT EXISTS user_sessions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        session_token TEXT UNIQUE,
        last_activity TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    
    # Create student academic records table
    cur.execute("""CREATE TABLE IF NOT EXISTS student_academic_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        qualification_type TEXT,
        institution_name TEXT,
        program_name TEXT,
        start_year INTEGER,
        end_year INTEGER,
        completion_status TEXT DEFAULT 'in_progress',
        certificate_file_url TEXT,
        verified INTEGER DEFAULT 0,
        verified_by_admin INTEGER DEFAULT 0,
        verification_notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    
    # Create semester results table
    cur.execute("""CREATE TABLE IF NOT EXISTS semester_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        semester_number INTEGER,
        year INTEGER,
        gpa REAL,
        courses TEXT,
        transcript_file_url TEXT,
        verified INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    
    # Create university programs table
    cur.execute("""CREATE TABLE IF NOT EXISTS university_programs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        faculty TEXT NOT NULL,
        program_name TEXT NOT NULL,
        program_type TEXT,
        entry_requirements TEXT,
        duration TEXT
    )""")

    # Create academic program course catalog table
    cur.execute("""CREATE TABLE IF NOT EXISTS program_courses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        faculty TEXT,
        program_name TEXT NOT NULL,
        semester INTEGER,
        course_name TEXT
    )""")

    # Create student course selections table
    cur.execute("""CREATE TABLE IF NOT EXISTS student_courses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        program_name TEXT NOT NULL,
        semester_number INTEGER,
        course_name TEXT,
        selected INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")

    # Create student portal import audit table
    cur.execute("""CREATE TABLE IF NOT EXISTS student_data_imports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        student_id TEXT NOT NULL,
        portal_url TEXT,
        data_json TEXT,
        status TEXT DEFAULT 'pending',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")

    # Ensure student course summary is persisted when selected
    existing_student_cols = [row[1] for row in conn.execute("PRAGMA table_info(students)").fetchall()]
    if 'selected_courses' not in existing_student_cols:
        conn.execute("ALTER TABLE students ADD COLUMN selected_courses TEXT")
    if 'anchor_tx' not in existing_student_cols:
        try:
            conn.execute("ALTER TABLE students ADD COLUMN anchor_tx TEXT")
        except Exception:
            pass

    # ========== CHAT SYSTEM TABLES ==========
    cur.execute("""CREATE TABLE IF NOT EXISTS chat_sessions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        chat_type TEXT DEFAULT 'ai',
        is_active INTEGER DEFAULT 1,
        admin_id INTEGER,
        escalation_reason TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        closed_at TIMESTAMP,
        FOREIGN KEY(user_id) REFERENCES users(id),
        FOREIGN KEY(admin_id) REFERENCES users(id)
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS chat_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id INTEGER NOT NULL,
        content TEXT NOT NULL,
        is_from_user INTEGER DEFAULT 1,
        sender_id INTEGER,
        receiver_id INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(session_id) REFERENCES chat_sessions(id),
        FOREIGN KEY(sender_id) REFERENCES users(id),
        FOREIGN KEY(receiver_id) REFERENCES users(id)
    )""")

    # ========== COMPANY SYSTEM TABLES ==========
    cur.execute("""CREATE TABLE IF NOT EXISTS companies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        name TEXT NOT NULL,
        description TEXT,
        website_url TEXT,
        linkedin_url TEXT,
        logo_url TEXT,
        industry TEXT,
        location TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(user_id) REFERENCES users(id)
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS company_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        data_type TEXT,
        fetched_data TEXT,
        last_fetched TIMESTAMP,
        fetch_status TEXT DEFAULT 'pending',
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        title TEXT NOT NULL,
        message TEXT NOT NULL,
        link TEXT,
        type TEXT DEFAULT 'system',
        icon TEXT DEFAULT 'fas fa-bell',
        read_status INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_notifications_user ON notifications(user_id, created_at DESC)")

    # ========== ADVANCED CHAT SYSTEM TABLES ==========
    cur.execute("""CREATE TABLE IF NOT EXISTS admin_forwarded_chats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        admin_id INTEGER,
        original_question TEXT NOT NULL,
        status TEXT DEFAULT 'pending',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        resolved_at TIMESTAMP,
        FOREIGN KEY(user_id) REFERENCES users(id),
        FOREIGN KEY(admin_id) REFERENCES users(id)
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS friend_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sender_id INTEGER NOT NULL,
        receiver_id INTEGER NOT NULL,
        status TEXT DEFAULT 'pending',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        responded_at TIMESTAMP,
        FOREIGN KEY(sender_id) REFERENCES users(id),
        FOREIGN KEY(receiver_id) REFERENCES users(id),
        UNIQUE(sender_id, receiver_id)
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS chat_blocks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        blocker_id INTEGER NOT NULL,
        blocked_id INTEGER NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(blocker_id) REFERENCES users(id),
        FOREIGN KEY(blocked_id) REFERENCES users(id),
        UNIQUE(blocker_id, blocked_id)
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS student_chat_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sender_id INTEGER NOT NULL,
        receiver_id INTEGER NOT NULL,
        content TEXT NOT NULL,
        is_read INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(sender_id) REFERENCES users(id),
        FOREIGN KEY(receiver_id) REFERENCES users(id)
    )""")

    # ========== ADD MISSING COLUMNS IF NEEDED ==========
    existing_chat_cols = [row[1] for row in conn.execute("PRAGMA table_info(chat_sessions)").fetchall()]
    # Additional columns can be added here if needed

    
    # Create indexes for better performance
    cur.execute("CREATE INDEX IF NOT EXISTS idx_messages_sender_receiver ON messages(sender_id, receiver_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_messages_created_at ON messages(created_at)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_job_interests_student ON job_interests(student_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_job_interests_job ON job_interests(job_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_users_role ON users(role)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_jobs_status ON jobs(status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_message_requests_sender ON message_requests(sender_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_message_requests_receiver ON message_requests(receiver_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_scraped_jobs_posted_date ON scraped_jobs(posted_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_results_verifications_student ON results_verifications(student_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_user_sessions_user_id ON user_sessions(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_user_sessions_last_activity ON user_sessions(last_activity)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_student_academic_records_user ON student_academic_records(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_semester_results_user ON semester_results(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_university_programs_faculty ON university_programs(faculty)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_program_courses_program ON program_courses(program_name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_student_courses_user ON student_courses(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_student_data_imports_user ON student_data_imports(user_id)")
    
    # Indexes for new chat system
    cur.execute("CREATE INDEX IF NOT EXISTS idx_admin_forwarded_chats_user ON admin_forwarded_chats(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_admin_forwarded_chats_admin ON admin_forwarded_chats(admin_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_admin_forwarded_chats_status ON admin_forwarded_chats(status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_friend_requests_sender ON friend_requests(sender_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_friend_requests_receiver ON friend_requests(receiver_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_friend_requests_status ON friend_requests(status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_chat_blocks_blocker ON chat_blocks(blocker_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_chat_blocks_blocked ON chat_blocks(blocked_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_student_chat_messages_sender ON student_chat_messages(sender_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_student_chat_messages_receiver ON student_chat_messages(receiver_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_student_chat_messages_created ON student_chat_messages(created_at)")
    
    # Indexes for chat system
    cur.execute("CREATE INDEX IF NOT EXISTS idx_chat_sessions_user ON chat_sessions(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_chat_sessions_admin ON chat_sessions(admin_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_chat_messages_session ON chat_messages(session_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_chat_messages_created_at ON chat_messages(created_at)")
    
    # Indexes for company system
    cur.execute("CREATE INDEX IF NOT EXISTS idx_companies_user ON companies(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_companies_name ON companies(name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_company_data_company ON company_data(company_id)")

    cur.execute("""CREATE TABLE IF NOT EXISTS scraped_company_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        scraped_description TEXT,
        scraped_logo_url TEXT,
        scraped_contact TEXT,
        social_links TEXT,
        last_scraped TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        is_valid INTEGER DEFAULT 1,
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS external_jobs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT,
        company TEXT,
        location TEXT,
        salary TEXT,
        rating REAL,
        url TEXT,
        source TEXT,
        scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS posted_jobs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        title TEXT NOT NULL,
        description TEXT NOT NULL,
        requirements TEXT,
        location TEXT,
        job_type TEXT,
        deadline TIMESTAMP,
        is_active INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )""")
    
    conn.commit()
    
    # Load university programs from prospectus
    programs_exist = conn.execute("SELECT COUNT(*) FROM university_programs").fetchone()[0] > 0
    if not programs_exist:
        programs = [
            # FACULTY OF DESIGN AND INNOVATION
            ('FACULTY OF DESIGN AND INNOVATION', 'Fashion and Retailing', 'degree', '', '3 years'),
            ('FACULTY OF DESIGN AND INNOVATION', 'Advertising', 'diploma', '', '2 years'),
            ('FACULTY OF DESIGN AND INNOVATION', 'Graphic Design', 'diploma', '', '2 years'),
            ('FACULTY OF DESIGN AND INNOVATION', 'Fashion and Apparel Design', 'diploma', '', '2 years'),
            
            # FACULTY OF COMMUNICATION, MEDIA AND BROADCASTING
            ('FACULTY OF COMMUNICATION, MEDIA AND BROADCASTING', 'Professional Communication', 'degree', '', '3 years'),
            ('FACULTY OF COMMUNICATION, MEDIA AND BROADCASTING', 'Broadcasting & Journalism', 'degree', '', '3 years'),
            ('FACULTY OF COMMUNICATION, MEDIA AND BROADCASTING', 'Digital Film & Television', 'degree', '', '3 years'),
            ('FACULTY OF COMMUNICATION, MEDIA AND BROADCASTING', 'Television and Film Production', 'diploma', '', '2 years'),
            ('FACULTY OF COMMUNICATION, MEDIA AND BROADCASTING', 'Broadcasting (Radio and TV)', 'diploma', '', '2 years'),
            ('FACULTY OF COMMUNICATION, MEDIA AND BROADCASTING', 'Public Relations', 'diploma', '', '2 years'),
            ('FACULTY OF COMMUNICATION, MEDIA AND BROADCASTING', 'Journalism and Media', 'diploma', '', '2 years'),
            
            # FACULTY OF ARCHITECTURE AND THE BUILT ENVIRONMENT
            ('FACULTY OF ARCHITECTURE AND THE BUILT ENVIRONMENT', 'Architectural Technology', 'diploma', '', '2 years'),
            
            # FACULTY OF BUSINESS AND GLOBALIZATION
            ('FACULTY OF BUSINESS AND GLOBALIZATION', 'International Business', 'degree', '', '3 years'),
            ('FACULTY OF BUSINESS AND GLOBALIZATION', 'Entrepreneurship', 'degree', '', '3 years'),
            ('FACULTY OF BUSINESS AND GLOBALIZATION', 'Human Resource Management', 'degree', '', '3 years'),
            ('FACULTY OF BUSINESS AND GLOBALIZATION', 'Business Management', 'diploma', '', '2 years'),
            ('FACULTY OF BUSINESS AND GLOBALIZATION', 'Retail Management', 'diploma', '', '2 years'),
            ('FACULTY OF BUSINESS AND GLOBALIZATION', 'Marketing', 'diploma', '', '2 years'),
            
            # FACULTY OF CREATIVITY IN TOURISM AND HOSPITALITY
            ('FACULTY OF CREATIVITY IN TOURISM AND HOSPITALITY', 'Tourism Management', 'degree', '', '3 years'),
            ('FACULTY OF CREATIVITY IN TOURISM AND HOSPITALITY', 'International Tourism', 'diploma', '', '2 years'),
            ('FACULTY OF CREATIVITY IN TOURISM AND HOSPITALITY', 'Tourism Management', 'diploma', '', '2 years'),
            ('FACULTY OF CREATIVITY IN TOURISM AND HOSPITALITY', 'Events Management', 'diploma', '', '2 years'),
            
            # FACULTY OF INFORMATION AND COMMUNICATION TECHNOLOGY
            ('FACULTY OF INFORMATION AND COMMUNICATION TECHNOLOGY', 'Software Engineering with Multimedia', 'degree', '', '3 years'),
            ('FACULTY OF INFORMATION AND COMMUNICATION TECHNOLOGY', 'Business Information Technology', 'degree', '', '3 years'),
            ('FACULTY OF INFORMATION AND COMMUNICATION TECHNOLOGY', 'Information Technology', 'degree', '', '3 years'),
            ('FACULTY OF INFORMATION AND COMMUNICATION TECHNOLOGY', 'Multimedia and Software Engineering', 'diploma', '', '2 years'),
            ('FACULTY OF INFORMATION AND COMMUNICATION TECHNOLOGY', 'Business Information Technology', 'diploma', '', '2 years'),
            ('FACULTY OF INFORMATION AND COMMUNICATION TECHNOLOGY', 'Information Technology', 'diploma', '', '2 years'),
        ]
        
        for faculty, program_name, program_type, requirements, duration in programs:
            conn.execute("""
                INSERT INTO university_programs (faculty, program_name, program_type, entry_requirements, duration)
                VALUES (?, ?, ?, ?, ?)
            """, (faculty, program_name, program_type, requirements, duration))
        conn.commit()
    
    conn.close()

    # Default Admin
    conn = get_db()
    cur = conn.cursor()
    if not cur.execute("SELECT 1 FROM users WHERE email = 'admin@portal.co.ls'").fetchone():
        cur.execute("INSERT INTO users (name, email, password_hash, role) VALUES (?,?,?,?)",
                    ("System Admin", "admin@portal.co.ls", generate_password_hash("admin123"), "admin"))
        conn.commit()
    conn.close()


def load_program_course_catalog():
    """Populate program_courses from course_data.json when the database is empty."""
    try:
        conn = get_db()
        cur = conn.cursor()
        existing = cur.execute("SELECT COUNT(*) FROM program_courses").fetchone()[0]
        if existing > 0:
            conn.close()
            return

        source_path = os.path.join(BASE_DIR, 'course_data.json')
        if not os.path.exists(source_path):
            conn.close()
            return

        with open(source_path, 'r', encoding='utf-8') as f:
            catalog = json.load(f)

        insert_query = """
            INSERT INTO program_courses (faculty, program_name, semester, course_name)
            VALUES (?, ?, ?, ?)
        """
        for program_name, program_data in catalog.items():
            faculty = program_data.get('faculty')
            semesters = program_data.get('semesters', {})
            for semester, courses in semesters.items():
                for course_name in courses:
                    cur.execute(insert_query, (faculty, program_name, int(semester), course_name))
        conn.commit()
    except Exception as e:
        print(f"Error loading program course catalog: {e}")
    finally:
        try:
            conn.close()
        except Exception:
            pass


init_db()
chatbot_engine = ChatbotEngine()
load_program_course_catalog()

# Initialize SQLAlchemy tables (safe: will create missing tables only)
try:
    init_models()
except Exception as e:
    print(f"SQLAlchemy init skipped or failed: {e}")


def fetch_from_university_portal(student_id, portal_url=None, password=None, audit_user_id=None):
    # Attempt to fetch real data from a student portal. Heuristics only; may require custom selectors for your portal.
    sid = (student_id or '').upper()
    if not portal_url:
        portal_url = os.environ.get('UNIVERSITY_PORTAL_URL')
    if not portal_url:
        print('Portal fetch aborted: no UNIVERSITY_PORTAL_URL configured')
        return None
    result = {
        'student_id': sid,
        'name': None,
        'course': None,
        'faculty': None,
        'gpa': None,
        'graduation_year': None
    }

    if not portal_url:
        return None

    # Basic URL and host validation to prevent SSRF and unsafe requests
    try:
        parsed = urlparse(portal_url)
        hostname = parsed.hostname
        if not hostname:
            print("Portal fetch aborted: invalid hostname")
            return None
        # Require HTTPS
        if parsed.scheme.lower() != 'https':
            print("Portal fetch aborted: only HTTPS endpoints are allowed")
            return None
    except Exception:
        print("Portal fetch aborted: invalid portal_url")
        return None

    # Allowlist hosts - tighten this list to trusted portals only
    ALLOWED_PORTAL_HOSTS = {"portal.co.ls", "www.portal.co.ls"}
    def hostname_allowed(h):
        try:
            h_l = h.lower()
            for a in ALLOWED_PORTAL_HOSTS:
                if h_l == a or h_l.endswith('.' + a):
                    return True
            return False
        except Exception:
            return False

    if not hostname_allowed(hostname):
        print(f"Portal fetch aborted: hostname '{hostname}' not in allowlist")
        return None

    # Resolve host and reject private/loopback/reserved addresses to prevent SSRF
    def resolves_to_private(h):
        try:
            infos = socket.getaddrinfo(h, None)
            for info in infos:
                ip = info[4][0]
                try:
                    ip_obj = ipaddress.ip_address(ip)
                    if ip_obj.is_private or ip_obj.is_loopback or ip_obj.is_multicast or ip_obj.is_reserved or ip_obj.is_link_local:
                        return True
                except Exception:
                    # if parsing fails, be conservative and treat as private
                    return True
            return False
        except Exception:
            # DNS resolution failed; treat as unsafe
            return True

    if resolves_to_private(hostname):
        print(f"Portal fetch aborted: hostname '{hostname}' resolves to private/reserved address")
        return None

    try:
        session_req = requests.Session()
        # reduce redirects and require TLS verification
        session_req.max_redirects = 5
        headers = {'User-Agent': 'Mozilla/5.0'}

        # First try: attempt a login POST if password provided
        login_tried = False
        if password:
            login_tried = True
            try:
                post_data = {'student_id': student_id, 'password': password}
                session_req.post(portal_url, data=post_data, headers=headers, timeout=10)
            except Exception:
                # try posting to /login
                try:
                    session_req.post(urljoin(portal_url, '/login'), data=post_data, headers=headers, timeout=10)
                except Exception:
                    pass

        # Then try to fetch a profile-like page
        profile_urls = [portal_url, urljoin(portal_url, '/profile'), urljoin(portal_url, f'/students/{sid}'), urljoin(portal_url, f'/student/{sid}')]
        html = None
        for pu in profile_urls:
            try:
                r = session_req.get(pu, headers=headers, timeout=10, allow_redirects=True, verify=True)
                # Reject very large responses to avoid memory exhaustion
                if r.status_code == 200 and len(r.content) > 100 and len(r.content) < 500000:
                    html = r.text
                    break
                else:
                    # continue to next candidate
                    continue
            except requests.exceptions.SSLError:
                print("Portal fetch aborted: SSL verification failed")
                return None
            except Exception:
                continue

        if not html:
            return None

        # If the portal returns JSON
        try:
            data_json = json.loads(html)
            # Map and whitelist expected keys only
            result['name'] = data_json.get('name') or data_json.get('fullName') or data_json.get('studentName')
            result['course'] = data_json.get('course') or data_json.get('program')
            result['faculty'] = data_json.get('faculty') or data_json.get('department')
            # Cast/validate numeric types
            try:
                result['gpa'] = float(data_json.get('gpa')) if data_json.get('gpa') is not None else None
            except Exception:
                result['gpa'] = None
            try:
                gy = data_json.get('graduation_year') or data_json.get('graduationYear')
                result['graduation_year'] = int(gy) if gy else None
            except Exception:
                result['graduation_year'] = None
            # Sanitize strings if bleach is available
            if bleach:
                for k in ['name', 'course', 'faculty']:
                    if result.get(k):
                        result[k] = bleach.clean(str(result[k]))
            return result
        except Exception:
            pass

        soup = BeautifulSoup(html, 'html.parser')

        # Heuristics: look for labels and nearby text
        text = soup.get_text(separator='|')
        text_l = text.lower()

        # Name
        name_tag = None
        for sel in ['#student-name', '.student-name', 'h1', 'h2', 'title']:
            t = soup.select_one(sel)
            if t and t.get_text(strip=True):
                name_tag = t.get_text(strip=True)
                break
        if not name_tag:
            # fallback: find line with student id and take previous/next token
            m = re.search(r"([A-Z]{1,2}\d{3,6})", text)
            if m:
                # try to find surrounding name
                parts = text.split('|')
                for i,p in enumerate(parts):
                    if sid in p:
                        if i>0:
                            name_tag = parts[i-1].strip()
                            break
        if name_tag:
            result['name'] = name_tag

        # Course/faculty/gpa/graduation by keyword
        # Course
        m = re.search(r'course[:\s]*([A-Za-z0-9 &/\-]+)', text, re.IGNORECASE)
        if m:
            result['course'] = m.group(1).strip()
        else:
            # try common labels
            for lbl in ['program', 'degree', 'course of study', 'study']:
                if lbl in text_l:
                    seg = re.search(rf"{lbl}[:\s]*([A-Za-z0-9 &/\-]+)", text, re.IGNORECASE)
                    if seg:
                        result['course'] = seg.group(1).strip()
                        break

        # Faculty/department
        m = re.search(r'(faculty|department)[:\s]*([A-Za-z &]+)', text, re.IGNORECASE)
        if m:
            result['faculty'] = m.group(2).strip()

        # GPA
        m = re.search(r'gpa[:\s]*([0-4]\.?\d{0,2})', text, re.IGNORECASE)
        if m:
            try:
                result['gpa'] = float(m.group(1))
            except Exception:
                pass

        # Graduation year
        m = re.search(r'(graduat(e|ion) year|class of)[:\s]*(20\d{2})', text, re.IGNORECASE)
        if m:
            try:
                result['graduation_year'] = int(m.group(3))
            except Exception:
                pass

        # Ensure at least a name or course found
        if not result['name'] and not result['course']:
            return None

        # Audit import into student_data_imports (avoid storing raw HTML)
        try:
            conn = get_db()
            cur = conn.cursor()
            audit_json = json.dumps(result)
            user_id = audit_user_id or (session.get('user_id') if session and session.get('user_id') else 0)
            cur.execute("INSERT INTO student_data_imports (user_id, student_id, portal_url, data_json, status) VALUES (?, ?, ?, ?, ?)",
                        (user_id, sid, portal_url, audit_json, 'fetched'))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Failed to write student_data_imports audit: {e}")

        return result
    except Exception as e:
        print(f"Portal fetch error: {e}")
        return None

# ===================== WEBSITE METADATA EXTRACTION =====================

def extract_website_metadata(company_url):
    """Extract website metadata from a company URL."""
    metadata = {
        'title': None,
        'description': None,
        'logo_url': None,
        'industry': None,
        'contact_email': None,
        'phone': None,
        'social_links': [],
        'recent_updates': []
    }
    try:
        response = requests.get(company_url, timeout=10, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        if response.status_code != 200:
            return metadata

        soup = BeautifulSoup(response.text, 'html.parser')

        if soup.title and soup.title.string:
            metadata['title'] = soup.title.string.strip()

        description_tag = soup.find('meta', attrs={'name': 'description'}) or soup.find('meta', attrs={'property': 'og:description'})
        if description_tag and description_tag.get('content'):
            metadata['description'] = description_tag['content'].strip()

        logo_tag = soup.find('meta', attrs={'property': 'og:image'}) or soup.find('link', attrs={'rel': 'shortcut icon'}) or soup.find('link', attrs={'rel': 'icon'})
        if logo_tag:
            url_value = logo_tag.get('content') or logo_tag.get('href')
            if url_value:
                metadata['logo_url'] = urljoin(company_url, url_value.strip())

        email_match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", response.text)
        if email_match:
            metadata['contact_email'] = email_match.group(0)

        phone_match = re.search(r"\+?\d[\d\s()\-]{7,}\d", response.text)
        if phone_match:
            metadata['phone'] = phone_match.group(0).strip()

        metadata['social_links'] = extract_social_links(company_url)

        recent_links = []
        for a in soup.find_all('a', href=True)[:30]:
            text = a.get_text(' ', strip=True)
            href = a['href'].strip()
            if any(keyword in (text.lower() + ' ' + href.lower()) for keyword in ['news', 'blog', 'update', 'press', 'story']):
                recent_links.append({'title': text or 'Latest update', 'url': urljoin(company_url, href)})
        metadata['recent_updates'] = recent_links[:5]

        # Try to infer company industry from keywords or structured data
        industry_tag = soup.find('meta', attrs={'name': 'industry'}) or soup.find('meta', attrs={'property': 'industry'})
        if industry_tag and industry_tag.get('content'):
            metadata['industry'] = industry_tag['content'].strip()
        else:
            text_body = (soup.get_text(separator=' ') or '').lower()
            industry_keywords = {
                'technology': ['software', 'tech', 'technology', 'digital', 'it', 'cloud', 'saas'],
                'finance': ['finance', 'bank', 'insurance', 'fintech', 'investment'],
                'education': ['education', 'school', 'university', 'learning', 'training'],
                'healthcare': ['health', 'medical', 'clinic', 'hospital', 'pharma'],
                'manufacturing': ['manufacturing', 'factory', 'industrial', 'production'],
                'consulting': ['consulting', 'business advisory', 'services', 'consultant'],
                'marketing': ['marketing', 'advertising', 'branding', 'creative'],
                'logistics': ['logistics', 'transport', 'shipping', 'supply chain']
            }
            for industry, keywords in industry_keywords.items():
                if any(keyword in text_body for keyword in keywords):
                    metadata['industry'] = industry.title()
                    break

        if not metadata['description']:
            ld_json = soup.find('script', type='application/ld+json')
            if ld_json and ld_json.string:
                try:
                    import json
                    data = json.loads(ld_json.string)
                    if isinstance(data, dict) and 'description' in data:
                        metadata['description'] = data['description'].strip()
                    elif isinstance(data, list):
                        for item in data:
                            if isinstance(item, dict) and 'description' in item:
                                metadata['description'] = item['description'].strip()
                                break
                except Exception:
                    pass
    except Exception:
        pass

    return metadata


def extract_social_links(company_url):
    """Extract social media links from a company website."""
    social_links = []
    try:
        response = requests.get(company_url, timeout=10, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        if response.status_code != 200:
            return social_links

        soup = BeautifulSoup(response.text, 'html.parser')
        domains = ['linkedin.com', 'facebook.com', 'twitter.com', 'instagram.com', 'youtube.com', 't.me', 'github.com']
        anchors = soup.find_all('a', href=True)
        found = set()
        for a in anchors:
            href = a['href'].strip()
            if any(domain in href for domain in domains):
                full_url = href
                if href.startswith('/'):
                    full_url = urljoin(company_url, href)
                if full_url not in found:
                    found.add(full_url)
                    social_links.append(full_url)
    except Exception:
        pass
    return social_links


def get_company_scrape_snapshot(company_id):
    """Return cached website metadata for a company, refreshing every 24 hours."""
    conn = get_db()
    company = conn.execute("SELECT * FROM companies WHERE id = ?", (company_id,)).fetchone()
    if not company:
        conn.close()
        return None

    latest = conn.execute("""
        SELECT * FROM company_data
        WHERE company_id = ?
        ORDER BY last_fetched DESC LIMIT 5
    """, (company_id,)).fetchall()

    needs_refresh = not latest or any(
        (datetime.utcnow() - row['last_fetched']).total_seconds() > 60 * 60 * 24
        for row in latest
    )

    snapshot = {
        'company': company,
        'website': None,
        'linkedin': None,
        'social_links': [],
        'recent_updates': [],
        'contact_email': None,
        'phone': None,
        'description': company['description'] or 'No description provided yet.',
        'logo_url': company['logo_url']
    }

    if needs_refresh and company['website_url']:
        try:
            website_metadata = extract_website_metadata(company['website_url'])
            snapshot.update({
                'website': website_metadata,
                'description': website_metadata.get('description') or snapshot['description'],
                'logo_url': website_metadata.get('logo_url') or snapshot['logo_url'],
                'contact_email': website_metadata.get('contact_email'),
                'phone': website_metadata.get('phone'),
                'social_links': website_metadata.get('social_links', []) or snapshot['social_links'],
                'recent_updates': website_metadata.get('recent_updates', []) or snapshot['recent_updates']
            })
            conn.execute("""
                INSERT OR REPLACE INTO company_data (company_id, data_type, fetched_data, last_fetched, fetch_status)
                VALUES (?, 'website', ?, CURRENT_TIMESTAMP, 'success')
            """, (company_id, json.dumps(website_metadata)))
        except Exception:
            pass

    for row in latest:
        try:
            data = json.loads(row['fetched_data']) if row['fetched_data'] else {}
            if row['data_type'] == 'website':
                snapshot['website'] = data
                if data.get('description'):
                    snapshot['description'] = data['description']
                if data.get('logo_url'):
                    snapshot['logo_url'] = data['logo_url']
                if data.get('contact_email'):
                    snapshot['contact_email'] = data['contact_email']
                if data.get('phone'):
                    snapshot['phone'] = data['phone']
                if data.get('social_links'):
                    snapshot['social_links'] = data['social_links']
                if data.get('recent_updates'):
                    snapshot['recent_updates'] = data['recent_updates']
        except Exception:
            continue

    conn.commit()
    conn.close()
    return snapshot


def create_notification_for_user(user_id, title, message, link=None, notification_type='system', icon='fas fa-bell'):
    if not user_id:
        return None
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO notifications (user_id, title, message, link, type, icon, read_status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, 0, CURRENT_TIMESTAMP)
    """, (user_id, title, message, link, notification_type, icon))
    conn.commit()
    conn.close()
    return True


def generate_recruiter_matches(recruiter_id, top_n=6):
    conn = get_db()
    recruiter = conn.execute("SELECT * FROM recruiters WHERE id = ?", (recruiter_id,)).fetchone()
    if not recruiter:
        conn.close()
        return []

    recruiter_keywords = set(re.findall(r"[A-Za-z0-9+#]+", ' '.join(filter(None, [
        recruiter['company_name'], recruiter['industry'], recruiter['description'], recruiter['website_title'], recruiter['website_description']
    ])) .lower()))

    student_rows = conn.execute("""
        SELECT u.id, u.name, u.email, s.course, s.faculty, s.gpa, s.skills, s.internship_status, s.results
        FROM users u
        JOIN students s ON s.user_id = u.id
        WHERE u.role IN ('student', 'alumni')
    """).fetchall()
    conn.close()

    ranked = []
    for student in student_rows:
        student_text = ' '.join(filter(None, [student['skills'], student['course'], student['faculty'], student['results'], student['internship_status']])) .lower()
        student_tokens = set(re.findall(r"[A-Za-z0-9+#]+", student_text))
        overlap = len(recruiter_keywords.intersection(student_tokens))
        try:
            gpa_value = float(student['gpa']) if student['gpa'] is not None else 0.0
        except (TypeError, ValueError):
            gpa_value = 0.0
        gpa_bonus = min(gpa_value / 4.0, 1.0) * 0.25
        status_bonus = 0.1 if (student['internship_status'] or '').lower() in ['seeking internship', 'open to work'] else 0
        score = min(1.0, (overlap / max(1, len(recruiter_keywords))) * 0.65 + gpa_bonus + status_bonus)
        if score >= 0.18:
            ranked.append({
                'user_id': student['id'],
                'name': student['name'],
                'email': student['email'],
                'course': student['course'],
                'faculty': student['faculty'],
                'gpa': student['gpa'],
                'skills': student['skills'],
                'score': round(score, 2)
            })

    ranked.sort(key=lambda item: (-item['score'], item['name']))
    return ranked[:top_n]


def process_recruiter_company_profile(recruiter_id=None, user_id=None):
    conn = get_db()
    recruiter = None
    if recruiter_id:
        recruiter = conn.execute("SELECT * FROM recruiters WHERE id = ?", (recruiter_id,)).fetchone()
    elif user_id:
        recruiter = conn.execute("SELECT * FROM recruiters WHERE user_id = ?", (user_id,)).fetchone()

    if not recruiter:
        conn.close()
        return None

    company_url = recruiter['company_website'] or recruiter['website']
    if not company_url:
        conn.execute("UPDATE recruiters SET scrape_status = 'failed', scrape_message = 'No company website configured', scrape_error = 'Missing website', last_scrape_at = CURRENT_TIMESTAMP WHERE id = ?", (recruiter['id'],))
        conn.commit()
        conn.close()
        return None

    try:
        conn.execute("UPDATE recruiters SET scrape_status = 'scraping', scrape_message = 'Refreshing company profile and job listings', scrape_error = NULL, last_scrape_at = CURRENT_TIMESTAMP WHERE id = ?", (recruiter['id'],))
        conn.commit()

        metadata = extract_website_metadata(company_url)
        social_links = extract_social_links(company_url)
        company_name = metadata.get('title') or recruiter['company_name'] or recruiter['website']
        description = metadata.get('description') or recruiter['description']
        logo_url = metadata.get('logo_url') or recruiter['logo_url'] or recruiter['website_logo_url']
        industry = metadata.get('industry') or recruiter['industry']

        conn.execute("""
            UPDATE recruiters
            SET company_name = ?, industry = ?, website_title = ?, description = ?, logo_url = ?, website_logo_url = ?, social_links = ?, company_website = ?,
                scrape_status = 'completed', scrape_message = 'Company profile refreshed', last_scrape_at = CURRENT_TIMESTAMP, scrape_error = NULL
            WHERE id = ?
        """, (company_name, industry, metadata.get('title'), description, logo_url, logo_url, ','.join(social_links), company_url, recruiter['id']))

        company = conn.execute("SELECT id FROM companies WHERE user_id = ?", (recruiter['user_id'],)).fetchone()
        company_payload = {
            'title': metadata.get('title'),
            'description': description,
            'logo_url': logo_url,
            'industry': industry,
            'contact_email': metadata.get('contact_email'),
            'phone': metadata.get('phone'),
            'social_links': social_links,
            'recent_updates': metadata.get('recent_updates', []),
            'website_url': company_url
        }
        if company:
            conn.execute("""
                UPDATE companies
                SET name = ?, description = ?, website_url = ?, linkedin_url = ?, logo_url = ?, industry = ?, location = ?, updated_at = ?
                WHERE id = ?
            """, (company_name, description, company_url, social_links[0] if social_links else None, logo_url, industry, recruiter['location'], datetime.now(), company['id']))
            company_id = company['id']
        else:
            conn.execute("""
                INSERT INTO companies (user_id, name, description, website_url, linkedin_url, logo_url, industry, location)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (recruiter['user_id'], company_name, description, company_url, social_links[0] if social_links else None, logo_url, industry, recruiter['location']))
            company_id = conn.lastrowid

        conn.execute("""
            INSERT OR REPLACE INTO company_data (company_id, data_type, fetched_data, last_fetched, fetch_status)
            VALUES (?, 'website', ?, CURRENT_TIMESTAMP, 'success')
        """, (company_id, json.dumps(company_payload)))

        jobs_found = scrape_jobs_from_url(company_url, recruiter['id'])
        conn.execute("UPDATE recruiters SET jobs_found_count = ? WHERE id = ?", (jobs_found, recruiter['id']))

        matches = generate_recruiter_matches(recruiter['id'])
        for match in matches:
            create_notification_for_user(
                match['user_id'],
                '🎯 Recruiter match alert',
                f"{company_name} is looking for students with {', '.join((match['skills'] or '').split(',')[:3]) or 'matching skills'}.",
                '/dashboard',
                'match',
                'fas fa-user-graduate'
            )

        conn.commit()
        return {'company': company_name, 'jobs_found': jobs_found, 'matches': len(matches)}
    except Exception as exc:
        conn.execute("UPDATE recruiters SET scrape_status = 'failed', scrape_message = 'Failed to refresh company profile', scrape_error = ? WHERE id = ?", (str(exc)[:250], recruiter['id']))
        conn.commit()
        conn.close()
        return None
    finally:
        conn.close()

# ===================== COMPANY VERIFICATION (AUTO) =====================

def verify_company_automatically(company_url, company_email=None):
    """
    Automatically verify company legitimacy using free public data
    Returns: (status, score, reasons_details)
    """
    score = 0
    reasons = []
    
    # 1. Check if website exists and has business info
    try:
        response = requests.get(company_url, timeout=10, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        if response.status_code == 200:
            score += 10
            reasons.append("✅ Website is accessible")
            
            soup = BeautifulSoup(response.text, 'html.parser')
            metadata = extract_website_metadata(company_url)
            if metadata['title']:
                score += 5
                reasons.append("✅ Website title captured")
            if metadata['description']:
                score += 5
                reasons.append("✅ Website description captured")
            if metadata['logo_url']:
                score += 5
                reasons.append("✅ Website image/logo found")
            
            # Look for business indicators
            business_terms = ['about', 'company', 'contact', 'address', 'phone', 
                            'email', 'business', 'office', 'terms', 'privacy']
            found_terms = sum(1 for term in business_terms 
                            if term in response.text.lower())
            
            if found_terms >= 3:
                score += 20
                reasons.append(f"✅ Found {found_terms} business indicators")
            
            # Check for contact page
            if soup.find('a', href=lambda x: x and 'contact' in x.lower()):
                score += 15
                reasons.append("✅ Contact page exists")
                
    except Exception as e:
        reasons.append(f"❌ Website check failed: {str(e)[:50]}")
    
    # 2. Check domain age
    domain = company_url.replace('http://', '').replace('https://', '').replace('www.', '').split('/')[0]
    if whois is not None:
        try:
            domain_info = whois.whois(domain)
            if domain_info.creation_date:
                if isinstance(domain_info.creation_date, list):
                    creation_date = domain_info.creation_date[0]
                else:
                    creation_date = domain_info.creation_date
                age_days = (datetime.now() - creation_date).days
                if age_days > 730:  # 2+ years
                    score += 25
                    reasons.append(f"✅ Domain age: {age_days//365} years (trusted)")
                elif age_days > 365:  # 1-2 years
                    score += 15
                    reasons.append(f"✅ Domain age: {age_days//365} year (acceptable)")
                else:
                    reasons.append(f"⚠️ New domain ({age_days} days old)")
        except Exception:
            reasons.append(f"⚠️ Could not verify domain age")
    else:
        reasons.append("⚠️ WHOIS lookup unavailable")

    # 3. Check email domain match
    if company_email:
        try:
            email_domain = company_email.split('@')[1]
            website_domain = domain.replace('www.', '')
            if email_domain == website_domain or email_domain in website_domain:
                score += 20
                reasons.append("✅ Email matches website domain (professional)")
            else:
                reasons.append("⚠️ Email uses free service (Gmail/Hotmail)")
        except Exception:
            pass

    # 4. Check DNS records
    if dns is not None:
        try:
            mx_records = dns.resolver.resolve(domain, 'MX')
            if mx_records:
                score += 10
                reasons.append("✅ Valid mail server configured")
        except Exception:
            reasons.append("⚠️ No dedicated mail server")
    else:
        reasons.append("⚠️ DNS resolver unavailable")
    
    # Determine status
    if score >= VERIFICATION_THRESHOLDS['auto_approve']:
        status = "auto_verified"
    elif score >= VERIFICATION_THRESHOLDS['manual_review']:
        status = "pending_review"
    else:
        status = "rejected"
    
    return {
        'status': status,
        'score': score,
        'reasons': reasons,
        'metadata': metadata if 'metadata' in locals() else {'title': None, 'description': None, 'logo_url': None},
        'needs_manual_review': score < VERIFICATION_THRESHOLDS['auto_approve']
    }


def extract_student_query_terms(incoming_message):
    text = (incoming_message or '').strip().lower()
    grade_match = re.search(r'\b([abcdf])\b', text, re.I)
    grade = grade_match.group(1).upper() if grade_match else None

    percent_match = re.search(r'(?:>|at least|minimum|more than|above|over)\s*(\d{1,3})\s*%', text, re.I)
    if percent_match:
        min_percent = int(percent_match.group(1))
    else:
        percent_match = re.search(r'\b(\d{1,3})\s*%\b', text)
        min_percent = int(percent_match.group(1)) if percent_match else None

    subject_match = re.search(r'\b(?:in|for|about|on)\s+([a-zA-Z][a-zA-Z\- ]{2,})', text)
    subject = subject_match.group(1).strip() if subject_match else None
    if subject and subject.lower().startswith('the '):
        subject = subject[4:]
    return {
        'grade': grade,
        'min_percent': min_percent,
        'subject': subject,
        'text': text,
    }


def query_students_for_assistant(question):
    terms = extract_student_query_terms(question)
    if not terms['subject'] and not terms['grade'] and terms['min_percent'] is None:
        return None

    from models import SessionLocal
    session = SessionLocal()
    try:
        sql = """
            SELECT u.name, u.email, s.course, s.faculty, s.gpa, s.results, s.skills
            FROM users u
            JOIN students s ON s.user_id = u.id
            WHERE u.role IN ('student','alumni')
        """
        params = {}
        clauses = []

        if terms['subject']:
            subject_like = f"%{terms['subject']}%"
            clauses.append("(LOWER(s.course) LIKE :subject OR LOWER(s.skills) LIKE :subject OR LOWER(s.results) LIKE :subject)")
            params['subject'] = subject_like.lower()

        if terms['grade']:
            clauses.append("(LOWER(s.results) LIKE :grade OR LOWER(s.skills) LIKE :grade)")
            params['grade'] = f"%{terms['grade'].lower()}%"

        if terms['min_percent'] is not None:
            clauses.append("(s.gpa >= :min_gpa)")
            params['min_gpa'] = float(terms['min_percent'])

        if clauses:
            sql += ' AND ' + ' AND '.join(clauses)

        sql += ' ORDER BY s.gpa DESC LIMIT 20'
        rows = session.execute(text(sql), params).fetchall()
        if not rows:
            return "I could not find matching student records for that query."

        lines = ["Here are the student records I found:"]
        for row in rows:
            lines.append(f"- {row.name} | Course: {row.course or 'N/A'} | Faculty: {row.faculty or 'N/A'} | GPA/Score: {row.gpa or 'N/A'}")
        return '\n'.join(lines)
    finally:
        session.close()


def generate_chatbot_response(incoming_message, user_id=None, session_id=None):
    try:
        result = get_ai_response(
            incoming_message,
            user_id=user_id,
            session_id=session_id or f"legacy_{user_id or 'anonymous'}"
        )
        return result["response"]
    except Exception as exc:
        logger.exception("Advanced chatbot processing failed: %s", exc)
        return "I’m here to help. Please try your question again."


def get_primary_admin_id():
    conn = get_db()
    admin = conn.execute("SELECT id FROM users WHERE role='admin' ORDER BY id LIMIT 1").fetchone()
    conn.close()
    return admin['id'] if admin else None

# ===================== CHAT MODERATION (AUTO) =====================

def is_job_related_text(text):
    clean_text = (text or "").strip().lower()
    if not clean_text:
        return False

    keywords = ['job', 'internship', 'vacancy', 'career', 'position', 'hiring', 'apply', 'salary', 'responsibilities', 'deadline']
    score = sum(1 for k in keywords if k in clean_text)
    if score >= 2:
        return True

    if ML_AVAILABLE and sentence_encoder and np is not None:
        try:
            prompt = 'Job posting summary, role description, responsibilities, company information, location and salary'
            vec_text = sentence_encoder.encode([clean_text])[0]
            vec_prompt = sentence_encoder.encode([prompt])[0]
            similarity = np.dot(vec_text, vec_prompt) / (np.linalg.norm(vec_text) * np.linalg.norm(vec_prompt))
            return similarity > 0.65
        except Exception:
            pass

    return False


def parse_job_card(card):
    text = ' '.join(card.stripped_strings)
    title = None
    company = None
    location = None
    salary = None
    posted_date = None

    if card.name == 'article' or card.name == 'div':
        for header_tag in ['h1', 'h2', 'h3', 'h4']:
            header = card.find(header_tag)
            if header and len(header.get_text(strip=True)) > 10:
                title = header.get_text(strip=True)
                break

    if not title:
        title = card.get_text(strip=True)[:120]

    if card.find(string=lambda s: s and 'company' in s.lower()):
        company_text = card.find(string=lambda s: s and 'company' in s.lower())
        company = company_text.parent.get_text(strip=True).replace('Company', '').strip()

    location_tag = card.find(string=lambda s: s and 'location' in s.lower())
    if location_tag:
        location = location_tag.parent.get_text(strip=True).replace('Location', '').strip()

    salary_tag = card.find(string=lambda s: s and 'salary' in s.lower())
    if salary_tag:
        salary = salary_tag.parent.get_text(strip=True).replace('Salary', '').strip()

    if card.find(string=lambda s: s and 'posted' in s.lower()):
        posted_date = card.find(string=lambda s: s and 'posted' in s.lower()).strip()

    return title, ' '.join(card.stripped_strings), company, location, salary, posted_date


def store_scraped_job(source_url, title, description, company=None, location=None, salary=None, job_type=None, posted_date=None):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        existing = cur.execute("SELECT 1 FROM scraped_jobs WHERE source_url = ? AND title = ?", (source_url, title)).fetchone()
        if existing:
            return False
        cur.execute("""
            INSERT INTO scraped_jobs (source_url, title, description, company, location, salary, job_type, posted_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (source_url, title, description, company, location, salary, job_type, posted_date))
        conn.commit()
        return True
    finally:
        if conn:
            conn.close()


def scrape_jobs_from_url(url, recruiter_id=None):
    jobs_found = 0
    try:
        response = requests.get(url, timeout=12, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        if response.status_code != 200:
            return jobs_found

        soup = BeautifulSoup(response.text, 'html.parser')
        containers = soup.find_all(['article', 'div', 'li'], limit=150)
        for card in containers:
            text = ' '.join(card.stripped_strings)
            if len(text) < 80:
                continue
            if not is_job_related_text(text):
                continue
            title, description, company, location, salary, posted_date = parse_job_card(card)
            if not title or len(title) < 8:
                continue
            if store_scraped_job(url, title, description, company, location, salary, None, posted_date):
                jobs_found += 1
    except requests.exceptions.ConnectionError as ce:
        safe_print(f"Connection error scraping {url}: DNS or network issue")
    except requests.exceptions.Timeout as te:
        safe_print(f"Timeout scraping {url}: Request took too long")
    except requests.exceptions.RequestException as re:
        safe_print(f"Request error scraping {url}: {re}")
    except Exception as e:
        safe_print(f"Scrape error for {url}: {e}")

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE recruiters SET last_scrape_at = ?, jobs_found_count = ? WHERE id = ?", (datetime.now(), jobs_found, recruiter_id))
        conn.commit()
    except Exception as update_error:
        safe_print(f"Failed to update recruiter scrape stats for {url}: {update_error}")
    finally:
        if conn:
            conn.close()

    return jobs_found


def refresh_scraped_jobs():
    conn = None
    try:
        conn = get_db()
        recruiter_rows = conn.execute("SELECT id, company_website, website FROM recruiters WHERE COALESCE(TRIM(company_website), TRIM(website)) != ''").fetchall()
    finally:
        if conn:
            conn.close()

    total = 0
    if recruiter_rows:
        for recruiter in recruiter_rows:
            result = process_recruiter_company_profile(recruiter_id=recruiter['id'])
            if result and result.get('jobs_found'):
                total += int(result['jobs_found'])
        safe_print(f"Refreshed {len(recruiter_rows)} recruiter company profiles and scraped {total} new jobs")
    else:
        sources = [
            'https://www.jobs.co.ls',
            'https://www.indeed.com/q-Lesotho-jobs.html',
            'https://www.countryjobs.com/lesotho-jobs',
            'https://www.glassdoor.com/Job/lesotho-jobs-SRCH_IL.0,7_IN103.htm'
        ]
        for source in sources:
            total += scrape_jobs_from_url(source)
        safe_print(f"Scraped {total} new jobs from online sources")

    # Keep only latest 300 entries
    conn = None
    try:
        conn = get_db()
        conn.execute("DELETE FROM scraped_jobs WHERE id NOT IN (SELECT id FROM scraped_jobs ORDER BY scraped_at DESC LIMIT 300)")
        conn.commit()
    finally:
        if conn:
            conn.close()


def sync_academic_data_for_user(user_id):
    conn = None
    try:
        conn = get_db()
        user = conn.execute("SELECT id, email, name, role, student_id, academic_data_retry_count FROM users WHERE id = ?", (user_id,)).fetchone()
        if not user:
            return False

        if not user['student_id']:
            conn.execute("UPDATE users SET academic_data_pending = 0, academic_data_last_attempt = ? WHERE id = ?", (datetime.utcnow(), user_id))
            conn.commit()
            return False

        start_attempt = datetime.utcnow()
        retry_count = (user['academic_data_retry_count'] or 0) + 1
        conn.execute("""
            UPDATE users
            SET academic_data_pending = 1,
                academic_data_last_attempt = ?,
                academic_data_retry_count = ?
            WHERE id = ?
        """, (start_attempt, retry_count, user_id))
        conn.commit()
    finally:
        if conn:
            conn.close()

    try:
        portal_data = fetch_from_university_portal(user['student_id'], audit_user_id=user_id)
    except Exception as e:
        print(f"Academic sync error for user {user_id}: {e}")
        portal_data = None

    conn = None
    try:
        conn = get_db()
        if portal_data:
            now = datetime.utcnow()
            portal_json = json.dumps(portal_data)
            conn.execute("""
                UPDATE users
                SET academic_data_pending = 0,
                    academic_data_last_attempt = ?,
                    academic_data_retry_count = ?,
                    academic_data_raw = ?
                WHERE id = ?
            """, (now, retry_count, portal_json, user_id))

            existing_student = conn.execute("SELECT id FROM students WHERE user_id = ?", (user_id,)).fetchone()
            if existing_student:
                conn.execute("""
                    UPDATE students
                    SET course = ?, faculty = ?, gpa = ?, graduation_year = ?
                    WHERE user_id = ?
                """, (portal_data.get('course'), portal_data.get('faculty'), portal_data.get('gpa'), portal_data.get('graduation_year'), user_id))
            else:
                default_status = "Not Seeking" if user['role'] == 'alumni' else "Seeking Internship"
                conn.execute("""
                    INSERT INTO students (user_id, course, faculty, gpa, graduation_year, internship_status)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (user_id, portal_data.get('course'), portal_data.get('faculty'), portal_data.get('gpa'), portal_data.get('graduation_year'), default_status))

            conn.commit()
        else:
            conn.execute("""
                UPDATE users
                SET academic_data_pending = 1,
                    academic_data_last_attempt = ?,
                    academic_data_retry_count = ?
                WHERE id = ?
            """, (datetime.utcnow(), retry_count, user_id))
            conn.commit()
    finally:
        if conn:
            conn.close()
    
    if portal_data:
        with app.app_context():
            send_email(
                user['email'],
                'Academic data synced',
                f"Your academic profile has been synced successfully. Course: {portal_data.get('course') or 'Not available'}"
            )
        return True
    
    return False


def retry_pending_academic_data_sync():
    conn = None
    try:
        conn = get_db()
        pending_users = conn.execute("SELECT id FROM users WHERE academic_data_pending = 1").fetchall()
    finally:
        if conn:
            conn.close()

    for pending_user in pending_users:
        threading.Thread(target=sync_academic_data_for_user, args=(pending_user['id'],), daemon=True).start()


def schedule_job_scraping():
    try:
        scheduler.add_job(refresh_scraped_jobs, 'interval', hours=6, id='refresh_scraped_jobs', replace_existing=True, next_run_time=datetime.now())
        scheduler.add_job(retry_pending_academic_data_sync, 'interval', hours=1, id='academic_data_retry_sync', replace_existing=True)
        scheduler.start()
        safe_print('Background schedulers started')
    except Exception as e:
        safe_print(f'Job scraping scheduler failed: {e}')

try:
    schedule_job_scraping()
except Exception as e:
    safe_print(f"Failed to schedule job scraping: {e}")

def generate_certificate_hash(user_id, results):
    """Generate a simple SHA-256 hash for a student's certificate/results."""
    if not results:
        return None
    base = f"{user_id}|{results}"
    return hashlib.sha256(base.encode('utf-8')).hexdigest()


def init_web3():
    try:
        provider_uri = os.environ.get('WEB3_PROVIDER_URI')
        private_key = os.environ.get('WEB3_PRIVATE_KEY')
        if not provider_uri or not private_key or Web3 is None:
            return None
        w3 = Web3(Web3.HTTPProvider(provider_uri))
        acct = w3.eth.account.from_key(private_key)
        return {'w3': w3, 'account': acct}
    except Exception as e:
        print(f"Web3 init error: {e}")
        return None


def anchor_certificate_on_chain(cert_hash):
    try:
        web3_ctx = init_web3()
        if not web3_ctx:
            return None
        w3 = web3_ctx['w3']
        acct = web3_ctx['account']
        to_addr = acct.address
        nonce = w3.eth.get_transaction_count(acct.address)
        tx = {
            'to': to_addr,
            'value': 0,
            'data': w3.toHex(text=cert_hash),
            'nonce': nonce,
            'gas': 21000,
            'gasPrice': w3.toWei(os.environ.get('WEB3_GAS_PRICE_GWEI', '5'), 'gwei')
        }
        signed = acct.sign_transaction(tx)
        tx_hash = w3.eth.send_raw_transaction(signed.rawTransaction)
        return tx_hash.hex()
    except Exception as e:
        print(f"Anchor error: {e}")
        return None

# Recommendation engine helpers
def build_job_profile_text(job):
    parts = [job.get('title'), job.get('description'), job.get('requirements'), job.get('skills_required'), job.get('location'), job.get('salary')]
    return ' '.join(str(part) for part in parts if part)

def build_student_profile_text(student):
    parts = [student.get('name'), student.get('course'), student.get('faculty'), student.get('skills'), student.get('results'), student.get('internship_status'), student.get('gpa'), student.get('graduation_year')]
    return ' '.join(str(part) for part in parts if part)

def compute_job_recommendations(job_id, top_n=10):
    conn = get_db()
    cur = conn.cursor()
    job = cur.execute("SELECT * FROM jobs WHERE id = ?", (job_id,)).fetchone()
    if not job:
        conn.close()
        return []

    students = cur.execute(
        """
        SELECT u.id AS user_id, u.name, s.course, s.faculty, s.gpa, s.graduation_year,
               s.skills, s.results, s.internship_status
        FROM users u
        JOIN students s ON u.id = s.user_id
        WHERE u.role IN ('student', 'alumni')
        """
    ).fetchall()

    if not students:
        conn.close()
        return []

    job_text = build_job_profile_text(job).strip()
    if not job_text:
        job_text = f"{job['title']} {job['location'] or ''} {job['salary'] or ''}"

    student_docs = []
    student_ids = []

    for student in students:
        student_ids.append(student['user_id'])
        student_docs.append(build_student_profile_text(student))

    if SKLEARN_AVAILABLE:
        try:
            vectorizer = TfidfVectorizer(stop_words='english', ngram_range=(1, 2))
            matrix = vectorizer.fit_transform([job_text] + student_docs)
            scores = cosine_similarity(matrix[0:1], matrix[1:]).flatten()
        except Exception as e:
            print(f"Recommendation engine error with sklearn: {e}")
            scores = None
    else:
        scores = None

    if scores is None:
        def score_text(source, target):
            source_tokens = set(re.findall(r"\w+", source.lower()))
            target_tokens = set(re.findall(r"\w+", target.lower()))
            overlap = source_tokens.intersection(target_tokens)
            return len(overlap) / max(1, len(source_tokens))

        scores = [score_text(job_text, doc) for doc in student_docs]

    recommendations = []
    for idx, student_id in enumerate(student_ids):
        recommendations.append({
            'student_id': student_id,
            'score': float(scores[idx])
        })

    if PANDAS_AVAILABLE:
        df = pd.DataFrame(recommendations)
        df = df.sort_values('score', ascending=False).head(top_n)
        recommendations = df.to_dict(orient='records')
    else:
        recommendations = sorted(recommendations, key=lambda x: x['score'], reverse=True)[:top_n]

    cur.execute("DELETE FROM job_recommendations WHERE job_id = ?", (job_id,))
    for rec in recommendations:
        cur.execute(
            "INSERT OR REPLACE INTO job_recommendations (job_id, student_id, score) VALUES (?, ?, ?)",
            (job_id, rec['student_id'], rec['score'])
        )
    conn.commit()
    conn.close()
    return recommendations

# End recommendation engine helpers

def moderate_message(message, sender_id=None, receiver_id=None):
    """Basic moderation: block messages containing banned words or suspicious patterns.
    Returns (is_allowed: bool, reason: str|None)
    """
    if not message:
        return True, None
    text = message.lower()
    for word in BANNED_WORDS:
        if word in text:
            log_moderation_flag(message, sender_id, receiver_id, f"banned_word:{word}")
            return False, f"Message contains banned word: {word}"
    for pat in SUSPICIOUS_PATTERNS:
        try:
            if re.search(pat, text):
                log_moderation_flag(message, sender_id, receiver_id, f"suspicious_pattern:{pat}")
                return False, "Message looks suspicious and was blocked"
        except re.error:
            continue
    return True, None

def update_user_session(user_id):
    """Update the user's session last_activity timestamp to keep 4-hour persistence."""
    try:
        conn = get_db()
        # Update last_activity timestamp - this keeps session alive for queries
        conn.execute("""
            UPDATE user_sessions SET last_activity = CURRENT_TIMESTAMP 
            WHERE user_id = ?
        """, (user_id,))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Session update error: {e}")

def get_or_create_session(user_id):
    """Get or create a user session token for 4-hour persistence."""
    try:
        conn = get_db()
        existing = conn.execute("""
            SELECT session_token FROM user_sessions 
            WHERE user_id = ? AND created_at > datetime('now', '-4 hours')
            ORDER BY created_at DESC LIMIT 1
        """, (user_id,)).fetchone()
        
        if existing:
            update_user_session(user_id)
            conn.close()
            return existing['session_token']
        
        # Create new session
        token = secrets.token_hex(32)
        conn.execute("""
            INSERT INTO user_sessions (user_id, session_token)
            VALUES (?, ?)
        """, (user_id, token))
        conn.commit()
        conn.close()
        return token
    except Exception as e:
        print(f"Session creation error: {e}")
        return None

def log_moderation_flag(message, sender_id, receiver_id, reason):
    """Log flagged messages for admin review"""
    conn = get_db()
    conn.execute("""
        INSERT INTO moderation_flags (message, sender_id, receiver_id, reason)
        VALUES (?, ?, ?, ?)
    """, (message, sender_id, receiver_id, reason))
    conn.commit()
    conn.close()

# ====================== ROUTES ======================

@app.route("/")
def home():
    return render_template("index.html")

@app.route('/course_data.json')
def course_data_json():
    return send_file(os.path.join(BASE_DIR, 'course_data.json'), mimetype='application/json')

@app.route("/register", methods=["GET", "POST"])
def register():
    if "user_id" in session:
        flash("You are already logged in", "info")
        return redirect(url_for("dashboard") if session.get("role") in ["student", "alumni"] else 
                       url_for("recruiter_dashboard") if session.get("role") == "recruiter" else 
                       url_for("admin_dashboard"))
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        password = request.form.get("password", "")
        role = request.form.get("role")
        student_number = normalize_student_number(request.form.get("student_id", ""))
        company_name = request.form.get("company_name", "").strip()
        company_website = request.form.get("company_website", "").strip()

        if role == "admin":
            flash("Admin accounts cannot be registered.", "danger")
            return render_template("register.html")

        if not all([name, email, password, role]):
            flash("Name, email, password and role are required", "danger")
            return render_template("register.html")

        if phone:
            try:
                import phonenumbers
                parsed_phone = phonenumbers.parse(phone, "LS")
                if not phonenumbers.is_valid_number(parsed_phone):
                    flash("Invalid phone number", "danger")
                    return render_template("register.html")
                phone = phonenumbers.format_number(parsed_phone, phonenumbers.PhoneNumberFormat.E164)
            except Exception:
                # Skip phone validation if phonenumbers not available
                pass
        else:
            phone = None

        if len(password) < 8:
            flash("Password must be at least 8 characters", "danger")
            return render_template("register.html")

        conn = get_db()
        cur = conn.cursor()

        if cur.execute("SELECT 1 FROM users WHERE email=?", (email,)).fetchone():
            flash("Email already registered", "warning")
            conn.close()
            return render_template("register.html")

        if role in ["student", "alumni"]:
            if not student_number:
                flash("Student Number is required for students and alumni", "danger")
                conn.close()
                return render_template("register.html")
            if not is_valid_student_number(student_number):
                flash("Student Number must contain exactly 9 digits.", "danger")
                conn.close()
                return render_template("register.html")
            existing_student_number = cur.execute(
                "SELECT id FROM users WHERE student_number = ? OR student_id = ?",
                (student_number, student_number)
            ).fetchone()
            if existing_student_number:
                flash("Student number already registered", "warning")
                conn.close()
                return render_template("register.html")

        if phone and cur.execute("SELECT 1 FROM users WHERE phone=?", (phone,)).fetchone():
            flash("Phone number already registered", "warning")
            conn.close()
            return render_template("register.html")

        pw_hash = generate_password_hash(password)

        try:
            cur.execute("""
                INSERT INTO users (name, email, phone, password_hash, role, student_id, student_number, academic_data_pending, academic_data_last_attempt, academic_data_retry_count, academic_data_raw)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (name, email, phone, pw_hash, role, student_number if role in ["student","alumni"] else None, student_number if role in ["student","alumni"] else None, 0, None, 0, None))
            user_id = cur.lastrowid

            if role in ["student", "alumni"]:
                if not student_number:
                    flash("Student Number is required", "danger")
                    conn.rollback()
                    conn.close()
                    return render_template("register.html")

                default_status = "Not Seeking" if role == "alumni" else "Seeking Internship"
                cur.execute("""
                    INSERT INTO students (user_id, course, faculty, gpa, graduation_year, internship_status)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (user_id, None, None, None, None, default_status))

                cur.execute("""
                    UPDATE users
                    SET academic_data_pending = 1,
                        academic_data_last_attempt = CURRENT_TIMESTAMP,
                        academic_data_retry_count = 0
                    WHERE id = ?
                """, (user_id,))

                conn.commit()
                threading.Thread(target=sync_academic_data_for_user, args=(user_id,), daemon=True).start()
                flash("Profile created. Academic data sync started in the background. You'll be notified when ready.", "info")

            elif role == "recruiter":
                if not company_website:
                    flash("Company Website URL is required for recruiters", "danger")
                    conn.rollback()
                    conn.close()
                    return render_template("register.html")

                cur.execute("""
                    INSERT INTO recruiters (user_id, company_name, website, company_website, verified, verification_score, verification_status, scrape_status, scrape_message)
                    VALUES (?, ?, ?, ?, 0, 0, 'pending', 'building', 'Analyzing company website')
                """, (user_id, company_name or name, company_website, company_website))
                recruiter_id = cur.lastrowid

                try:
                    metadata = extract_website_metadata(company_website)
                    if not metadata:
                        metadata = {'title': None, 'description': None, 'industry': None, 'logo_url': None}
                except Exception as meta_error:
                    print(f"Metadata extraction error: {meta_error}")
                    metadata = {'title': None, 'description': None, 'industry': None, 'logo_url': None}
                
                try:
                    social_links = extract_social_links(company_website)
                    social_links_text = ','.join(social_links) if social_links else None
                except Exception as social_error:
                    print(f"Social links extraction error: {social_error}")
                    social_links_text = None
                
                cur.execute("""
                    UPDATE recruiters
                    SET website_title = ?, description = ?, industry = ?, logo_url = ?, social_links = ?, company_website = ?, last_scrape_at = NULL, jobs_found_count = 0,
                        scrape_status = 'building', scrape_message = 'Building company intelligence profile'
                    WHERE id = ?
                """, (metadata.get('title'), metadata.get('description'), metadata.get('industry'), metadata.get('logo_url'), social_links_text, company_website, recruiter_id))
                threading.Thread(target=process_recruiter_company_profile, args=(recruiter_id,), daemon=True).start()
                flash("Recruiter profile created. AI company intelligence is running in the background.", "info")

            conn.commit()
            flash("Registration successful! Please login.", "success")
            return redirect(url_for("login"))
            
        except ValueError as ve:
            conn.rollback()
            print(f"Validation error during registration: {ve}")
            flash(f"Validation error: {str(ve)[:100]}", "danger")
            conn.close()
            return render_template("register.html")
        except TypeError as te:
            conn.rollback()
            print(f"Type error during registration: {te}")
            flash(f"Data format error. Please check your input.", "danger")
            conn.close()
            return render_template("register.html")
        except Exception as e:
            conn.rollback()
            print(f"Unexpected registration error: {e}")
            flash(f"Registration failed. Please try again.", "danger")
            conn.close()
            return render_template("register.html")

    return render_template("register.html")
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        flash("You are already logged in", "info")
        return redirect(url_for("dashboard") if session.get("role") in ["student", "alumni"] else 
                       url_for("recruiter_dashboard") if session.get("role") == "recruiter" else 
                       url_for("admin_dashboard"))
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
        conn.close()

        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            session["name"] = user["name"]
            session["profile_picture"] = user["profile_picture"] if user["profile_picture"] else "default-avatar.png"
            flash(f"Welcome, {user['name']}!", "success")

            if user["role"] == "admin":
                return redirect(url_for("admin_dashboard"))
            elif user["role"] == "recruiter":
                return redirect(url_for("recruiter_dashboard"))
            else:
                return redirect(url_for("dashboard"))
        flash("Invalid email or password", "danger")

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out", "info")
    return redirect(url_for("home"))

# Student / Alumni Dashboard
@app.route("/dashboard")
def dashboard():
    if session.get("role") not in ["student", "alumni"]:
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    profile = conn.execute("SELECT * FROM students WHERE user_id=?", (session["user_id"],)).fetchone()
    user = conn.execute("SELECT academic_data_pending FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    academic_data_sync_pending = bool(user['academic_data_pending']) if user else False
    jobs = conn.execute("""
        SELECT j.*, r.company_name, r.id AS company_id
        FROM jobs j
        JOIN recruiters r ON r.id = j.recruiter_id
        WHERE j.status = 'approved'
        ORDER BY j.created_at DESC
        LIMIT 6
    """).fetchall()
    recent_posted_jobs = conn.execute("""
        SELECT j.*, r.company_name, r.id AS company_id
        FROM jobs j
        JOIN recruiters r ON r.id = j.recruiter_id
        WHERE j.status = 'approved'
        ORDER BY j.created_at DESC
        LIMIT 5
    """).fetchall()
    
    # Get applications with status
    applications = conn.execute("""
        SELECT j.id as job_id, j.title, ji.created_at as applied_at,
               ji.status, ji.feedback
        FROM job_interests ji
        JOIN jobs j ON ji.job_id = j.id
        WHERE ji.student_id = ?
        ORDER BY ji.created_at DESC
    """, (session["user_id"],)).fetchall()
    
    scraped_jobs = conn.execute("""SELECT * FROM scraped_jobs WHERE status = 'active' ORDER BY scraped_at DESC LIMIT 10""").fetchall()

    skill_tokens = []
    if profile and profile['skills']:
        skill_tokens = [token.strip().lower() for token in str(profile['skills']).replace(';', ',').split(',') if token.strip()]

    recommended_jobs = []
    for job in jobs:
        job_text = ' '.join(filter(None, [job['title'], job['description'], job['skills_required'], job['requirements']])).lower()
        overlap = len(set(skill_tokens).intersection(set(re.findall(r"[A-Za-z0-9+#]+", job_text))))
        score = round(overlap / max(1, len(skill_tokens)) if skill_tokens else 0, 2)
        if overlap > 0 or score > 0:
            recommended_jobs.append({
                'job_id': job['id'],
                'title': job['title'],
                'company_name': job['company_name'],
                'score': score,
                'location': job['location'],
                'salary': job['salary'],
                'company_id': job['company_id']
            })

    recommended_jobs.sort(key=lambda item: (-item['score'], item['title']))
    completion_fields = [profile['course'] if profile else None, profile['faculty'] if profile else None, profile['gpa'] if profile else None, profile['skills'] if profile else None, profile['internship_status'] if profile else None]
    completion_percentage = int(round((sum(1 for field in completion_fields if field) / len(completion_fields)) * 100)) if completion_fields else 0
    conn.close()
    return render_template(
        "dashboard.html",
        profile=profile,
        jobs=jobs,
        applications=applications,
        scraped_jobs=scraped_jobs,
        recent_posted_jobs=recent_posted_jobs,
        academic_data_sync_pending=academic_data_sync_pending,
        recommended_jobs=recommended_jobs[:5],
        completion_percentage=completion_percentage,
        accepted_apps=sum(1 for app in applications if app['status'] == 'accepted'),
        pending_apps=sum(1 for app in applications if app['status'] == 'pending'),
        total_applications=len(applications)
    )

# Edit Profile
@app.route("/edit_profile", methods=["GET", "POST"])
def edit_profile():
    if session.get("role") not in ["student", "alumni"]:
        return redirect(url_for("home"))

    conn = get_db()
    profile = conn.execute("SELECT * FROM students WHERE user_id=?", (session["user_id"],)).fetchone()
    user = conn.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
    selected_courses = [row["course_name"] for row in conn.execute(
        "SELECT course_name FROM student_courses WHERE user_id=? ORDER BY semester_number, course_name", (session["user_id"],)
    ).fetchall()]
    import_history = conn.execute(
        "SELECT * FROM student_data_imports WHERE user_id=? ORDER BY created_at DESC LIMIT 5", (session["user_id"],)
    ).fetchall()

    if request.method == "POST":
        skills = request.form.get("skills", "").strip()
        internship_status = request.form.get("internship_status")
        course = request.form.get("course", "").strip() or None
        faculty = request.form.get("faculty", "").strip() or None
        phone = request.form.get("phone", "").strip() or None

        if profile:
            conn.execute(
                "UPDATE students SET course=?, faculty=?, skills=?, internship_status=? WHERE user_id=?",
                (course, faculty, skills, internship_status, session["user_id"])
            )
        else:
            conn.execute(
                "INSERT INTO students (user_id, course, faculty, skills, internship_status) VALUES (?,?,?,?,?)",
                (session["user_id"], course, faculty, skills, internship_status)
            )

        if phone:
            conn.execute("UPDATE users SET phone = ? WHERE id = ?", (phone, session['user_id']))

        conn.commit()
        flash("Profile saved", "success")
        conn.close()
        return redirect(url_for("edit_profile"))

    conn.close()
    return render_template("edit_profile.html", profile=profile, user=user, selected_courses=selected_courses, import_history=import_history)

@app.route("/upload_cv", methods=["POST"])
def upload_cv():
    if not session.get("user_id") or session.get("role") not in ["student", "alumni"]:
        flash("Unauthorized", "danger")
        return redirect(url_for("login"))

    if 'cv' not in request.files:
        flash("No file selected", "warning")
        return redirect(url_for("edit_profile"))

    file = request.files['cv']
    if file.filename == '' or not allowed_file(file.filename):
        flash('Invalid file. Allowed: PDF, DOC, DOCX, JPG, JPEG, PNG.', 'danger')
        return redirect(url_for('edit_profile'))

    filename = secure_filename(f"{session['user_id']}_{file.filename}")
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

    conn = get_db()
    existing = conn.execute("SELECT 1 FROM students WHERE user_id = ?", (session['user_id'],)).fetchone()
    if existing:
        conn.execute("UPDATE students SET cv_filename = ? WHERE user_id = ?", (filename, session['user_id']))
    else:
        conn.execute("INSERT INTO students (user_id, cv_filename) VALUES (?, ?)", (session['user_id'], filename))
    conn.commit()
    conn.close()

    flash('CV uploaded successfully.', 'success')
    return redirect(url_for('edit_profile'))

@app.route("/api/fetch_student_data", methods=["POST"])
def api_fetch_student_data():
    if session.get("role") not in ["student", "alumni"]:
        return jsonify({"error": "Unauthorized"}), 403

    student_id = request.form.get("student_id", "").strip().upper()
    portal_url = request.form.get("portal_url", "").strip()
    portal_password = request.form.get("portal_password", "").strip()

    if not student_id or not portal_url or not portal_password:
        return jsonify({"error": "Student ID, portal URL, and password are required."}), 400

    student_data = fetch_from_university_portal(student_id, portal_url, portal_password)
    if not student_data:
        return jsonify({"error": "Could not import student data from the portal."}), 404

    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO student_data_imports (user_id, student_id, portal_url, data_json, status) VALUES (?, ?, ?, ?, ?)",
            (session["user_id"], student_id, portal_url, json.dumps(student_data), "completed")
        )

        existing = cur.execute("SELECT 1 FROM students WHERE user_id=?", (session["user_id"],)).fetchone()
        internship_status = "Seeking Internship" if session.get("role") == "student" else "Not Seeking"
        if existing:
            cur.execute(
                "UPDATE students SET course=?, faculty=?, gpa=?, graduation_year=?, internship_status=? WHERE user_id=?",
                (student_data.get("course"), student_data.get("faculty"), student_data.get("gpa"), student_data.get("graduation_year"), internship_status, session["user_id"])
            )
        else:
            cur.execute(
                "INSERT INTO students (user_id, course, faculty, gpa, graduation_year, internship_status, skills) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (session["user_id"], student_data.get("course"), student_data.get("faculty"), student_data.get("gpa"), student_data.get("graduation_year"), internship_status, "")
            )

        cur.execute("UPDATE users SET student_id = ? WHERE id = ?", (student_id, session["user_id"]))
        conn.commit()
        conn.close()
        return jsonify({"success": True, "message": "Student portal data imported successfully.", "student_data": student_data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/save_selected_courses", methods=["POST"])
def save_selected_courses():
    if session.get("role") not in ["student", "alumni"]:
        return jsonify({"error": "Unauthorized"}), 403

    program_name = request.form.get("program_name", "").strip()
    semester = request.form.get("semester", "").strip()
    selected_courses = request.form.getlist("selected_courses[]") or request.form.getlist("selected_courses")

    if not program_name or not semester:
        return jsonify({"error": "Program name and semester are required."}), 400

    try:
        semester_number = int(semester)
    except ValueError:
        return jsonify({"error": "Semester must be a number."}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM student_courses WHERE user_id=? AND program_name=? AND semester_number=?",
        (session["user_id"], program_name, semester_number)
    )

    for course_name in selected_courses:
        cur.execute(
            "INSERT INTO student_courses (user_id, program_name, semester_number, course_name, selected) VALUES (?, ?, ?, ?, 1)",
            (session["user_id"], program_name, semester_number, course_name)
        )

    cur.execute(
        "UPDATE students SET selected_courses = ? WHERE user_id = ?",
        (','.join(selected_courses), session["user_id"])
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "selected_count": len(selected_courses)})

# Student Submit Results for Verification
@app.route("/submit_results_for_verification", methods=["POST"])
def submit_results_for_verification():
    if not session.get("user_id") or session.get("role") not in ["student", "alumni"]:
        return jsonify({"error": "Unauthorized"}), 403

    result_type = request.form.get("result_type", "Other").strip()
    results_text = request.form.get("results_text", "").strip()
    results_file = request.files.get("results_file")

    if not result_type or (result_type == 'Select'):
        return jsonify({"error": "Please choose a result type"}), 400
    if not results_text and not results_file:
        return jsonify({"error": "Please provide either text or file"}), 400

    try:
        conn = get_db()
        cur = conn.cursor()

        results_filename = None
        if results_file and results_file.filename:
            if allowed_file(results_file.filename):
                filename = secure_filename(f"{session['user_id']}_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{results_file.filename.rsplit('.', 1)[1].lower()}")
                results_file.save(os.path.join(UPLOAD_FOLDER, filename))
                results_filename = filename
            else:
                conn.close()
                return jsonify({"error": "File type not allowed"}), 400

        # Generate certificate hash
        cert_hash = generate_certificate_hash(session["user_id"], results_text or results_filename or "submitted")

        # Insert verification record
        cur.execute(
            """
            INSERT INTO results_verifications (student_id, result_type, results_text, results_file, certificate_hash, status)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (session["user_id"], result_type, results_text, results_filename, cert_hash, 'pending')
        )
        conn.commit()

        # Attempt automatic verification against SIS records
        sis_row = cur.execute(
            "SELECT s.* FROM sis_students s JOIN users u ON u.student_id = s.student_id WHERE u.id = ?",
            (session["user_id"],)
        ).fetchone()

        status = 'rejected'
        admin_notes = ''
        if sis_row:
            matched = False
            # Heuristics: match graduation year, GPA, or course name in submitted text
            try:
                sis_gpa = sis_row['gpa'] if 'gpa' in sis_row.keys() else None
            except Exception:
                sis_gpa = None
            sis_course = sis_row['course'] if 'course' in sis_row.keys() else None
            sis_grad = sis_row['graduation_year'] if 'graduation_year' in sis_row.keys() else None

            text = (results_text or "").lower()
            if sis_grad and str(sis_grad) in text:
                matched = True
            if sis_gpa is not None and (str(sis_gpa) in text or ("{:.1f}".format(sis_gpa) in text)):
                matched = True
            if sis_course and sis_course.lower() in text:
                matched = True

            if matched:
                status = 'verified'
                admin_notes = 'Auto-verified via SIS record match'
                # Update students record to mark verified and store certificate hash
                cur.execute("SELECT id FROM students WHERE user_id = ?", (session['user_id'],))
                stud = cur.fetchone()
                if stud:
                    # Anchor certificate on-chain if configured
                    anchor_tx = None
                    try:
                        anchor_tx = anchor_certificate_on_chain(cert_hash)
                    except Exception:
                        anchor_tx = None
                    cur.execute("UPDATE students SET certificate_hash = ?, certificate_verified = 1, anchor_tx = ? WHERE user_id = ?", (cert_hash, anchor_tx, session['user_id']))
                else:
                    anchor_tx = None
                    try:
                        anchor_tx = anchor_certificate_on_chain(cert_hash)
                    except Exception:
                        anchor_tx = None
                    cur.execute("INSERT INTO students (user_id, certificate_hash, certificate_verified, anchor_tx) VALUES (?,?,?,?)", (session['user_id'], cert_hash, 1, anchor_tx))
            else:
                status = 'rejected'
                admin_notes = 'Auto-rejected: submission did not match SIS records'
        else:
            status = 'rejected'
            admin_notes = 'Auto-rejected: no SIS record found for this account'

        # Update verification record with result (and anchor tx if available)
        try:
            # fetch anchor_tx from students table if set
            anchor_row = cur.execute("SELECT anchor_tx FROM students WHERE user_id = ?", (session['user_id'],)).fetchone()
            anchor_tx_val = anchor_row['anchor_tx'] if anchor_row and 'anchor_tx' in anchor_row.keys() else None
        except Exception:
            anchor_tx_val = None
        cur.execute(
            "UPDATE results_verifications SET status = ?, admin_notes = ?, verified_at = CURRENT_TIMESTAMP, anchor_tx = ? WHERE certificate_hash = ?",
            (status, admin_notes, anchor_tx_val, cert_hash)
        )
        conn.commit()

        # Notify user via SocketIO
        try:
            socketio.emit('verification_update', {'user_id': session['user_id'], 'status': status, 'message': admin_notes})
        except Exception:
            pass

        conn.close()
        return jsonify({"success": True, "message": "Results submitted and processed", "status": status})
    except Exception as e:
        print(f"Error submitting results: {e}")
        return jsonify({"error": str(e)}), 500

# Profile Picture Upload
@app.route("/upload_profile_picture", methods=["POST"])
def upload_profile_picture():
    if not session.get("user_id"):
        flash("Please login", "danger")
        return redirect(url_for("login"))
    
    if session.get("role") not in ['student', 'alumni']:
        flash("Only students and alumni can upload profile pictures.", "danger")
        return redirect(request.referrer)
    
    if 'profile_pic' not in request.files:
        flash("No file selected", "danger")
        return redirect(request.referrer)
    
    file = request.files['profile_pic']
    if file.filename == '':
        flash("No file selected", "danger")
        return redirect(request.referrer)
    
    if file and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = secure_filename(f"user_{session['user_id']}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
        file.save(os.path.join(app.config['PROFILE_PIC_FOLDER'], filename))
        conn = get_db()
        conn.execute("UPDATE users SET profile_picture = ? WHERE id = ?", (filename, session["user_id"]))
        conn.commit()
        conn.close()
        session["profile_picture"] = filename
        flash("Profile picture updated!", "success")
    else:
        flash("Invalid file type. Allowed: jpg, jpeg, png", "danger")
    
    return redirect(request.referrer)

# Recruiter Dashboard
@app.route("/recruiter_dashboard")
def recruiter_dashboard():
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    recruiter = conn.execute("SELECT * FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    jobs = conn.execute("SELECT * FROM jobs WHERE recruiter_id=(SELECT id FROM recruiters WHERE user_id=?)", (session["user_id"],)).fetchall()
    student_count = conn.execute("SELECT COUNT(*) FROM users u JOIN students s ON u.id=s.user_id WHERE u.role IN ('student','alumni')").fetchone()[0]
    available_students = conn.execute("""
        SELECT u.id, u.name, u.email, s.course, s.faculty, s.gpa, s.skills, s.certificate_hash
        FROM users u
        JOIN students s ON u.id = s.user_id
        WHERE u.role IN ('student','alumni')
        ORDER BY u.name
        LIMIT 6
    """).fetchall()
    recent_activity = conn.execute("""
        SELECT title, message, created_at
        FROM notifications
        WHERE user_id = ?
        ORDER BY created_at DESC
        LIMIT 6
    """, (session['user_id'],)).fetchall()
    candidate_matches = generate_recruiter_matches(recruiter['id']) if recruiter else []
    total_applications = conn.execute("SELECT COUNT(*) FROM job_interests ji JOIN jobs j ON ji.job_id = j.id WHERE j.recruiter_id = ?", (recruiter['id'],)).fetchone()[0] if recruiter else 0
    approved_jobs = conn.execute("SELECT COUNT(*) FROM jobs WHERE recruiter_id = ? AND status='approved'", (recruiter['id'],)).fetchone()[0] if recruiter else 0
    total_jobs = len(jobs)
    active_jobs = conn.execute("SELECT COUNT(*) FROM jobs WHERE recruiter_id = ? AND is_active = 1 AND is_filled = 0", (recruiter['id'],)).fetchone()[0] if recruiter else 0
    filled_jobs = conn.execute("SELECT COUNT(*) FROM jobs WHERE recruiter_id = ? AND is_filled = 1", (recruiter['id'],)).fetchone()[0] if recruiter else 0
    closed_jobs = conn.execute("SELECT COUNT(*) FROM jobs WHERE recruiter_id = ? AND is_active = 0 AND is_filled = 0", (recruiter['id'],)).fetchone()[0] if recruiter else 0
    profile_completion = round((sum(1 for field in [recruiter['company_name'], recruiter['website'], recruiter['description'], recruiter['website_logo_url'], recruiter['industry']] if field) / 5) * 100) if recruiter else 0
    conn.close()
    return render_template(
        "recruiter_dashboard.html",
        recruiter=recruiter,
        jobs=jobs,
        student_count=student_count,
        available_students=available_students,
        candidate_matches=candidate_matches,
        recent_activity=recent_activity,
        total_applications=total_applications,
        approved_jobs=approved_jobs,
        total_jobs=total_jobs,
        active_jobs=active_jobs,
        filled_jobs=filled_jobs,
        closed_jobs=closed_jobs,
        profile_completion=profile_completion
    )

@app.route("/recruiter/<int:recruiter_id>")
def recruiter_profile(recruiter_id):
    """View recruiter profile (job poster)"""
    if not session.get("user_id"):
        flash("Please log in to view recruiter profiles.", "warning")
        return redirect(url_for("login"))

    conn = get_db()
    recruiter = conn.execute(
        "SELECT r.*, u.email, u.name FROM recruiters r JOIN users u ON r.user_id=u.id WHERE r.id=?",
        (recruiter_id,)
    ).fetchone()
    conn.close()

    if not recruiter:
        flash("Recruiter profile not found.", "danger")
        return redirect(url_for("home"))

    return render_template("recruiter_profile.html", recruiter=recruiter)

# Post Job
@app.route("/post_job", methods=["GET", "POST"])
def post_job():
    if session.get("role") != "recruiter":
        flash("Only recruiters can post jobs", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    verified = conn.execute("SELECT verified FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    conn.close()

    if not verified or verified["verified"] == 0:
        flash("Your recruiter account is still pending admin approval. You cannot post jobs yet.", "warning")
        return redirect(url_for("recruiter_dashboard"))

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        requirements = request.form.get("requirements", "").strip()
        skills_required = request.form.get("skills_required", "").strip()
        location = request.form.get("location", "").strip()
        salary = request.form.get("salary", "").strip()

        if not title:
            flash("Job title is required", "danger")
            return render_template("post_job.html")

        conn = get_db()
        cur = conn.cursor()

        recruiter = cur.execute("SELECT id FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
        if not recruiter:
            flash("Recruiter profile not found", "danger")
            conn.close()
            return redirect(url_for("recruiter_dashboard"))

        cur.execute("""
            INSERT INTO jobs (recruiter_id, title, description, requirements, skills_required, location, salary, status, is_active, is_filled, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', 1, 0, ?)
        """, (recruiter["id"], title, description, requirements, skills_required, location, salary, datetime.now()))

        job_id = cur.lastrowid
        conn.commit()

        # Compute recommendations immediately for the new job
        try:
            compute_job_recommendations(job_id, top_n=10)
        except Exception as e:
            print(f"Recommendation generation failed: {e}")

        conn.close()

        flash("Job posted successfully (pending admin approval)", "success")
        return redirect(url_for("recruiter_dashboard"))

    return render_template("post_job.html")


@app.route("/recruiter/job/<int:job_id>/edit", methods=["GET", "POST"])
def edit_job(job_id):
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    recruiter = conn.execute("SELECT id FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    job = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    if not recruiter or not job or job["recruiter_id"] != recruiter["id"]:
        conn.close()
        flash("Job not found or access denied.", "danger")
        return redirect(url_for("recruiter_dashboard"))

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        requirements = request.form.get("requirements", "").strip()
        skills_required = request.form.get("skills_required", "").strip()
        location = request.form.get("location", "").strip()
        salary = request.form.get("salary", "").strip()

        if not title:
            flash("Job title is required", "danger")
            conn.close()
            return render_template("post_job.html", job=job)

        conn.execute("""
            UPDATE jobs
            SET title = ?, description = ?, requirements = ?, skills_required = ?, location = ?, salary = ?, updated_at = ?
            WHERE id = ?
        """, (title, description, requirements, skills_required, location, salary, datetime.now(), job_id))
        conn.commit()
        conn.close()
        flash("Job updated successfully.", "success")
        return redirect(url_for("recruiter_dashboard"))

    conn.close()
    return render_template("post_job.html", job=job)


@app.route("/recruiter/job/<int:job_id>/delete", methods=["POST"])
def delete_job(job_id):
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    recruiter = conn.execute("SELECT id FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    job = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    if recruiter and job and job["recruiter_id"] == recruiter["id"]:
        conn.execute("DELETE FROM jobs WHERE id = ?", (job_id,))
        conn.commit()
        flash("Job deleted successfully.", "success")
    else:
        flash("Unable to delete job.", "danger")
    conn.close()
    return redirect(url_for("recruiter_dashboard"))


@app.route("/recruiter/job/<int:job_id>/mark_filled", methods=["POST"])
def mark_job_filled(job_id):
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    recruiter = conn.execute("SELECT id FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    job = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    if recruiter and job and job["recruiter_id"] == recruiter["id"]:
        conn.execute("""
            UPDATE jobs
            SET is_filled = 1, is_active = 0, updated_at = ?
            WHERE id = ?
        """, (datetime.now(), job_id))
        conn.commit()
        flash("Job marked as filled.", "success")
    else:
        flash("Unable to update job status.", "danger")
    conn.close()
    return redirect(url_for("recruiter_dashboard"))


@app.route("/recruiter/job/<int:job_id>/mark_closed", methods=["POST"])
def mark_job_closed(job_id):
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    recruiter = conn.execute("SELECT id FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    job = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    if recruiter and job and job["recruiter_id"] == recruiter["id"]:
        conn.execute("""
            UPDATE jobs
            SET is_active = 0, updated_at = ?
            WHERE id = ?
        """, (datetime.now(), job_id))
        conn.commit()
        flash("Job marked as closed.", "success")
    else:
        flash("Unable to update job status.", "danger")
    conn.close()
    return redirect(url_for("recruiter_dashboard"))


# Recruiter view for job recommendations
@app.route("/recruiter/job/<int:job_id>/recommendations")
def job_recommendations(job_id):
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    recruiter = conn.execute("SELECT id FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    job = conn.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()

    if not recruiter or not job or job["recruiter_id"] != recruiter["id"]:
        conn.close()
        flash("Access denied", "danger")
        return redirect(url_for("recruiter_dashboard"))

    recommendations = conn.execute(
        """
        SELECT jr.score, u.id as user_id, u.name, u.email, s.course, s.faculty, s.gpa, s.skills, s.certificate_hash, s.internship_status
        FROM job_recommendations jr
        JOIN users u ON jr.student_id = u.id
        JOIN students s ON u.id = s.user_id
        WHERE jr.job_id = ?
        ORDER BY jr.score DESC
        LIMIT 20
        """,
        (job_id,)
    ).fetchall()

    if not recommendations:
        try:
            compute_job_recommendations(job_id, top_n=20)
            recommendations = conn.execute(
                """
                SELECT jr.score, u.id as user_id, u.name, s.course, s.faculty, s.gpa, s.skills, s.certificate_hash, s.internship_status
                FROM job_recommendations jr
                JOIN users u ON jr.student_id = u.id
                JOIN students s ON u.id = s.user_id
                WHERE jr.job_id = ?
                ORDER BY jr.score DESC
                LIMIT 20
                """,
                (job_id,)
            ).fetchall()
        except Exception as e:
            print(f"Recommendation refresh failed: {e}")

    conn.close()
    return render_template("job_recommendations.html", job=job, recommendations=recommendations)

# Search Students
@app.route("/search_students", methods=["GET", "POST"])
def search_students():
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    verified_row = conn.execute("SELECT verified FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    verified = verified_row["verified"] if verified_row else 0

    results = []
    limited = True

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        min_gpa = request.form.get("min_gpa")
        course = request.form.get("course", "").strip()
        skills = request.form.get("skills", "").strip()

        query = """
            SELECT u.id, u.name, s.course, s.faculty, s.gpa, s.skills, s.cv_filename,
                   s.results, s.certificate_hash,
                   s.graduation_year, s.internship_status
            FROM users u 
            JOIN students s ON u.id = s.user_id 
            WHERE 1=1
        """
        params = []

        if name:
            query += " AND u.name LIKE ?"
            params.append("%" + name + "%")
        if min_gpa:
            query += " AND s.gpa >= ?"
            params.append(float(min_gpa))
        if course:
            query += " AND s.course LIKE ?"
            params.append("%" + course + "%")
        if skills:
            query += " AND s.skills LIKE ?"
            params.append("%" + skills + "%")

        if verified == 1:
            limited = False
        else:
            query = """
                SELECT u.id, u.name, s.course, s.faculty 
                FROM users u 
                JOIN students s ON u.id = s.user_id
                WHERE 1=1
            """
            params = []
            if name:
                query += " AND u.name LIKE ?"
                params.append("%" + name + "%")
            if course:
                query += " AND s.course LIKE ?"
                params.append("%" + course + "%")

        results = conn.execute(query, params).fetchall()

    conn.close()
    return render_template("search_students.html", results=results, limited=limited)

# Recruiter Alumni View
@app.route("/recruiter/alumni")
def recruiter_alumni():
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))
    
    conn = get_db()
    alumni = conn.execute("""
        SELECT u.id, u.name, u.email, s.course, s.faculty, s.gpa, s.skills, s.cv_filename, s.results, s.certificate_hash
        FROM users u
        JOIN students s ON u.id = s.user_id
        WHERE u.role IN ('student', 'alumni')
        ORDER BY s.faculty, u.name
    """).fetchall()
    
    # Group by faculty
    faculty_order = ['FICT', 'FABE', 'FBMG', 'FCMB', 'FCTH', 'FDI', 'Other']
    faculty_groups = {fac: [] for fac in faculty_order}
    for a in alumni:
        faculty = a['faculty'] if a['faculty'] in faculty_order else 'Other'
        faculty_groups[faculty].append(a)
    faculty_groups = {k: v for k, v in faculty_groups.items() if v}
    
    conn.close()
    return render_template("recruiter_alumni.html", faculty_groups=faculty_groups)

@app.route("/student/<int:user_id>")
def student_detail(user_id):
    if not session.get("user_id"):
        flash("Please log in to view student profiles.", "warning")
        return redirect(url_for("login"))

    conn = get_db()
    student = conn.execute("SELECT u.id, u.name, u.email, s.course, s.faculty, s.gpa, s.skills, s.cv_filename, s.graduation_year, s.internship_status FROM users u JOIN students s ON u.id = s.user_id WHERE u.id = ?", (user_id,)).fetchone()
    conn.close()
    if not student:
        flash("Student not found.", "danger")
        return redirect(url_for("home"))

    return render_template("student_detail.html", student=student)

@app.route("/student/<int:student_id>/projects")
def student_projects(student_id):
    if not session.get("user_id"):
        flash("Please log in to view student projects.", "warning")
        return redirect(url_for("login"))

    conn = get_db()
    student = conn.execute("SELECT u.id, u.name, s.course, s.faculty FROM users u JOIN students s ON u.id = s.user_id WHERE u.id = ?", (student_id,)).fetchone()
    if not student:
        conn.close()
        flash("Student not found.", "danger")
        return redirect(url_for("home"))

    projects = conn.execute("SELECT id, title, description, link, created_at FROM projects WHERE user_id = ? ORDER BY created_at DESC", (student_id,)).fetchall()
    conn.close()
    return render_template("student_projects.html", student=student, projects=projects)

# Recruiter views interested students for their job
@app.route("/recruiter/job/<int:job_id>/interested")
def job_interested_students(job_id):
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    job = conn.execute("SELECT recruiter_id, title FROM jobs WHERE id=?", (job_id,)).fetchone()
    recruiter_id = conn.execute("SELECT id FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    
    if not job or job["recruiter_id"] != recruiter_id["id"]:
        flash("Access denied", "danger")
        conn.close()
        return redirect(url_for("recruiter_dashboard"))

    interested = conn.execute("""
        SELECT u.name, u.email, s.course, s.faculty, s.gpa, s.cv_filename, u.id as user_id,
               ji.id as application_id, ji.status, ji.feedback
        FROM job_interests ji
        JOIN users u ON ji.student_id = u.id
        JOIN students s ON u.id = s.user_id
        WHERE ji.job_id = ?
        ORDER BY ji.created_at DESC
    """, (job_id,)).fetchall()

    conn.close()
    return render_template("job_interested.html", interested=interested, job_id=job_id, job_title=job['title'])

# Update Application Status
@app.route("/recruiter/update_application/<int:application_id>", methods=["POST"])
def update_application_status(application_id):
    if session.get("role") != "recruiter":
        flash("Access denied", "danger")
        return redirect(url_for("home"))
    
    new_status = request.form.get("status")
    feedback = request.form.get("feedback", "")
    
    conn = get_db()
    app_data = conn.execute("""
        SELECT ji.*, j.recruiter_id, j.title, u.email as student_email, u.name as student_name
        FROM job_interests ji
        JOIN jobs j ON ji.job_id = j.id
        JOIN users u ON ji.student_id = u.id
        WHERE ji.id = ?
    """, (application_id,)).fetchone()
    
    recruiter_id = conn.execute("SELECT id FROM recruiters WHERE user_id=?", (session["user_id"],)).fetchone()
    
    if not app_data or app_data["recruiter_id"] != recruiter_id["id"]:
        flash("Unauthorized", "danger")
        conn.close()
        return redirect(url_for("recruiter_dashboard"))
    
    conn.execute("""
        UPDATE job_interests SET status = ?, feedback = ?, updated_at = ? WHERE id = ?
    """, (new_status, feedback, datetime.now(), application_id))
    conn.commit()
    
    # Send email notification
    send_email(app_data["student_email"],
               f"Application Status Update: {app_data['title']}",
               f"Dear {app_data['student_name']},\n\nYour application for '{app_data['title']}' has been {new_status}.\n\nFeedback: {feedback if feedback else 'No additional feedback provided.'}\n\nLogin to your dashboard for more details.")
    
    conn.close()
    flash(f"Application marked as {new_status} and student notified.", "success")
    return redirect(url_for("job_interested_students", job_id=app_data["job_id"]))

# Admin Dashboard
@app.route("/admin_dashboard")
def admin_dashboard():
    if session.get("role") != "admin":
        flash("Access denied - Admin only", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    total_companies = conn.execute("SELECT COUNT(*) FROM recruiters").fetchone()[0]
    total_jobs = conn.execute("SELECT COUNT(*) FROM jobs WHERE status='approved'").fetchone()[0]

    students = conn.execute("""
        SELECT u.id, u.name, u.email, u.role, s.course, s.faculty, s.gpa, s.skills, s.cv_filename
        FROM users u 
        LEFT JOIN students s ON u.id = s.user_id 
        WHERE u.role IN ('student', 'alumni')
        ORDER BY s.faculty, u.name
    """).fetchall()
    
    faculty_order = ['FICT', 'FABE', 'FBMG', 'FCMB', 'FCTH', 'FDI', 'Other']
    students_by_faculty = {fac: [] for fac in faculty_order}
    for s in students:
        faculty = s['faculty'] if s['faculty'] in faculty_order else 'Other'
        students_by_faculty[faculty].append(s)

    recruiters = conn.execute("""
        SELECT r.*, u.name, u.email 
        FROM recruiters r 
        JOIN users u ON r.user_id = u.id 
        ORDER BY u.name
    """).fetchall()
    
    
    jobs = conn.execute("""
        SELECT j.*, r.company_name, u.name as recruiter_name 
        FROM jobs j 
        JOIN recruiters r ON j.recruiter_id = r.id 
        JOIN users u ON r.user_id = u.id 
        ORDER BY j.id DESC
    """).fetchall()

    pending_results = conn.execute("SELECT rv.*, u.name, u.email, s.course FROM results_verifications rv JOIN users u ON rv.student_id = u.id JOIN students s ON u.id = s.user_id WHERE rv.status = 'pending' ORDER BY rv.created_at DESC").fetchall()
    pending_results_count = conn.execute("SELECT COUNT(*) FROM results_verifications WHERE status = 'pending'").fetchone()[0]

    broadcasts = conn.execute("""
        SELECT bm.*, u.name as admin_name, u.email as admin_email
        FROM broadcast_messages bm
        JOIN users u ON u.id = bm.admin_id
        ORDER BY bm.created_at DESC
        LIMIT 10
    """).fetchall()

    escalated_conversations = conn.execute("""
        SELECT cc.*, u.name as user_name, u.email as user_email
        FROM chatbot_conversations cc
        JOIN users u ON cc.user_id = u.id
        WHERE cc.escalated_to_admin = 1
        ORDER BY cc.created_at DESC
        LIMIT 20
    """).fetchall()
    escalated_count = conn.execute("SELECT COUNT(*) FROM chatbot_conversations WHERE escalated_to_admin = 1").fetchone()[0]
    
    # Monitoring data
    flagged = conn.execute("""
        SELECT mf.*, u1.name as sender, u2.name as receiver
        FROM moderation_flags mf
        JOIN users u1 ON mf.sender_id = u1.id
        JOIN users u2 ON mf.receiver_id = u2.id
        WHERE mf.reviewed = 0
        ORDER BY mf.created_at DESC
        LIMIT 100
    """).fetchall()
    
    recent_chats = conn.execute("""
        SELECT 
            m.*,
            u1.name as sender_name,
            u1.email as sender_email,
            u1.role as sender_role,
            u2.name as receiver_name,
            u2.email as receiver_email,
            u2.role as receiver_role
        FROM messages m
        JOIN users u1 ON m.sender_id = u1.id
        JOIN users u2 ON m.receiver_id = u2.id
        ORDER BY m.created_at DESC
        LIMIT 200
    """).fetchall()
    
    suspicious_users = conn.execute("""
        SELECT 
            u.id, u.name, u.email, u.role,
            COUNT(DISTINCT m.receiver_id) as unique_contacts,
            COUNT(m.id) as total_messages,
            AVG(LENGTH(m.message)) as avg_length,
            MAX(m.created_at) as last_active
        FROM users u
        LEFT JOIN messages m ON m.sender_id = u.id
        GROUP BY u.id
        HAVING unique_contacts > 30 OR total_messages > 200
        ORDER BY unique_contacts DESC
    """).fetchall()
    
    users_with_phones = conn.execute("""
        SELECT name, phone, role
        FROM users 
        WHERE phone IS NOT NULL AND phone != ''
        ORDER BY name
    """).fetchall()
    
    conn.close()

    return render_template("admin_dashboard.html", 
                         total_users=total_users,
                         total_students=total_students,
                         total_companies=total_companies,
                         total_jobs=total_jobs,
                         students_by_faculty=students_by_faculty,
                         recruiters=recruiters, 
                         jobs=jobs,
                         pending_results=pending_results,
                         pending_results_count=pending_results_count,
                         broadcasts=broadcasts,
                         escalated_conversations=escalated_conversations,
                         escalated_count=escalated_count,
                         flagged=flagged,
                         recent_chats=recent_chats,
                         suspicious_users=suspicious_users,
                         users_with_phones=users_with_phones)

# ===================== ADMIN RESULTS VERIFICATION =====================

@app.route("/admin/pending_results")
def admin_pending_results():
    """Get pending results verifications for admin review"""
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    conn = get_db()
    pending = conn.execute("""
        SELECT rv.*, u.name, u.email, s.course
        FROM results_verifications rv
        JOIN users u ON rv.student_id = u.id
        LEFT JOIN students s ON u.id = s.user_id
        WHERE rv.status = 'pending'
        ORDER BY rv.created_at DESC
    """).fetchall()
    
    conn.close()
    return jsonify([dict(row) for row in pending])

@app.route("/admin/verify_results_page")
def verify_results_page():
    """Admin results verification page"""
    if session.get("role") != "admin":
        flash("Unauthorized", "danger")
        return redirect(url_for("login"))
    
    return render_template("verify_results.html")

@app.route("/admin/verify_results/<int:verification_id>", methods=["POST"])
def admin_verify_results(verification_id):
    """Admin verifies or rejects student results"""
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    data = request.get_json()
    action = data.get("action")  # 'approve' or 'reject'
    notes = data.get("notes", "").strip()
    
    if action not in ["approve", "reject"]:
        return jsonify({"error": "Invalid action"}), 400
    
    conn = get_db()
    
    # Get the verification record
    verification = conn.execute("""
        SELECT * FROM results_verifications WHERE id = ?
    """, (verification_id,)).fetchone()
    
    if not verification:
        conn.close()
        return jsonify({"error": "Verification not found"}), 404
    
    # Update verification status
    status = "verified" if action == "approve" else "rejected"
    conn.execute("""
        UPDATE results_verifications
        SET status = ?, admin_notes = ?, verified_by = ?, verified_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (status, notes, session["user_id"], verification_id))
    
    # If approved, update student profile and create alert message
    if action == "approve":
        conn.execute("""
            UPDATE students
            SET certificate_verified = 1, certificate_hash = ?
            WHERE user_id = ?
        """, (verification["certificate_hash"], verification["student_id"]))
        
        # Insert admin alert message
        admin_user = conn.execute("SELECT id FROM users WHERE role = 'admin' AND email = 'admin@portal.co.ls'").fetchone()
        student_user = conn.execute("SELECT name FROM users WHERE id = ?", (verification["student_id"],)).fetchone()
        
        if admin_user and student_user:
            alert_msg = f"✅ Results Verified: {student_user['name']}'s results/certificate have been verified and are blockchain-certified."
            # Store as an admin broadcast/alert rather than a pinned message in the admin inbox
            try:
                conn.execute("""
                    INSERT INTO broadcast_messages (admin_id, message, sent_to_all)
                    VALUES (?, ?, 0)
                """, (admin_user['id'], alert_msg))
            except Exception:
                # Fallback to messages if broadcast table missing
                conn.execute("""
                    INSERT INTO messages (sender_id, receiver_id, message)
                    VALUES (?, ?, ?)
                """, (admin_user['id'], admin_user['id'], alert_msg))
    else:
        # If rejected, create alert message
        admin_user = conn.execute("SELECT id FROM users WHERE role = 'admin' AND email = 'admin@portal.co.ls'").fetchone()
        student_user = conn.execute("SELECT name FROM users WHERE id = ?", (verification["student_id"],)).fetchone()
        
        if admin_user and student_user:
            alert_msg = f"⚠️ Results Rejected: {student_user['name']}'s results were flagged as potentially forged/fake. Reason: {notes}"
            # Use broadcast_messages so alerts aren't pinned in the normal inbox
            try:
                conn.execute("""
                    INSERT INTO broadcast_messages (admin_id, message, sent_to_all)
                    VALUES (?, ?, 0)
                """, (admin_user['id'], alert_msg))
            except Exception:
                conn.execute("""
                    INSERT INTO messages (sender_id, receiver_id, message)
                    VALUES (?, ?, ?)
                """, (admin_user['id'], admin_user['id'], alert_msg))
    
    conn.commit()
    conn.close()
    
    return jsonify({"success": True, "message": f"Results {status} successfully"})

@app.route("/results_status/<int:student_id>")
def results_status(student_id):
    """Get results verification status for a student"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 403
    
    conn = get_db()
    
    # Check if user is the student or admin
    if session.get("role") != "admin" and session["user_id"] != student_id:
        conn.close()
        return jsonify({"error": "Unauthorized"}), 403
    
    verification = conn.execute("""
        SELECT * FROM results_verifications
        WHERE student_id = ?
        ORDER BY created_at DESC
        LIMIT 1
    """, (student_id,)).fetchone()
    
    conn.close()
    
    if not verification:
        return jsonify({"status": "none", "message": "No results submitted yet"})
    
    # Include student anchor tx if available
    v = dict(verification)
    try:
        if v.get('anchor_tx'):
            v['anchor_tx'] = v['anchor_tx']
    except Exception:
        pass
    return jsonify(v)

@app.route("/admin/send_bulk_email", methods=["POST"])
def admin_send_bulk_email():
    if session.get("role") != "admin":
        flash("Access denied", "danger")
        return redirect(url_for("home"))
    
    communication_method = request.form.get("communication_method", "email")
    recipient_group = request.form.get("recipient_group")
    subject = request.form.get("subject")
    message = request.form.get("message")
    
    if not subject or not message:
        flash("Subject and message are required", "danger")
        return redirect(url_for("admin_dashboard"))
    
    conn = get_db()
    if recipient_group == "all_students":
        recipients = conn.execute("SELECT id, email, role FROM users WHERE role IN ('student', 'alumni') AND id != ?", (session['user_id'],)).fetchall()
    elif recipient_group == "all_users":
        recipients = conn.execute("SELECT id, email, role FROM users WHERE id != ?", (session['user_id'],)).fetchall()
    elif recipient_group == "recruiters":
        recipients = conn.execute("SELECT u.id, u.email, u.role FROM users u JOIN recruiters r ON u.id = r.user_id WHERE u.id != ?", (session['user_id'],)).fetchall()
    else:
        flash("Invalid recipient group", "danger")
        conn.close()
        return redirect(url_for("admin_dashboard"))
    conn.close()
    
    sent_count = 0
    for recipient in recipients:
        try:
            if communication_method == 'chat':
                conn = get_db()
                conn.execute("INSERT INTO messages (sender_id, receiver_id, message) VALUES (?, ?, ?)",
                             (session['user_id'], recipient['id'], message))
                conn.commit()
                conn.close()
                sent_count += 1
            else:
                send_email(recipient['email'], subject, message)
                sent_count += 1
        except Exception as e:
            print(f"Failed to send to {recipient['email']}: {e}")
    
    if communication_method == 'chat':
        flash(f"Chat message sent to {sent_count} recipients", "success")
    else:
        flash(f"Email sent to {sent_count} recipients", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/analytics")
def admin_analytics():
    if session.get("role") != "admin":
        flash("Access denied - Admin only", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    total_students = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    total_companies = conn.execute("SELECT COUNT(*) FROM recruiters").fetchone()[0]
    total_jobs = conn.execute("SELECT COUNT(*) FROM jobs WHERE status='approved'").fetchone()[0]
    programs = conn.execute("SELECT COALESCE(course, 'Other') as course, COUNT(*) as count FROM students GROUP BY course ORDER BY count DESC").fetchall()
    program_names = [row["course"] for row in programs]
    program_counts = [row["count"] for row in programs]
    employed = conn.execute("SELECT COUNT(*) FROM students WHERE internship_status != 'Seeking Internship'").fetchone()[0]
    seeking = conn.execute("SELECT COUNT(*) FROM students WHERE internship_status = 'Seeking Internship'").fetchone()[0]
    conn.close()

    return render_template("analytics.html",
                         total_users=total_users,
                         total_students=total_students,
                         total_companies=total_companies,
                         total_jobs=total_jobs,
                         programs=program_names,
                         program_counts=program_counts,
                         employed=employed,
                         seeking=seeking)

# Approve Job
@app.route("/approve_job/<int:job_id>")
def approve_job(job_id):
    if session.get("role") != "admin":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    conn.execute("UPDATE jobs SET status='approved' WHERE id=?", (job_id,))
    conn.commit()
    
    job = conn.execute("SELECT recruiter_id, title FROM jobs WHERE id=?", (job_id,)).fetchone()
    recruiter = conn.execute("SELECT u.email FROM users u JOIN recruiters r ON r.user_id = u.id WHERE r.id=?", (job['recruiter_id'],)).fetchone()
    
    if recruiter:
        send_email(recruiter['email'], f"Job Approved: {job['title']}", "Your job has been approved and is now visible.")
    
    conn.close()
    flash("Job approved and now visible to students/alumni", "success")
    return redirect(url_for("admin_dashboard"))

# View Job Details
@app.route("/admin/view_job/<int:job_id>")
def admin_view_job(job_id):
    if session.get("role") != "admin":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    job = conn.execute("""
        SELECT j.*, r.company_name, u.name as recruiter_name 
        FROM jobs j 
        JOIN recruiters r ON j.recruiter_id = r.id 
        JOIN users u ON r.user_id = u.id 
        WHERE j.id = ?
    """, (job_id,)).fetchone()
    conn.close()

    return render_template("admin_view_job.html", job=job)

# Verify Recruiter
@app.route("/verify_recruiter/<int:user_id>")
def verify_recruiter(user_id):
    if session.get("role") != "admin":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    conn.execute("UPDATE recruiters SET verified=1, verification_status='admin_verified' WHERE user_id=?", (user_id,))
    conn.commit()
    
    recruiter = conn.execute("SELECT email FROM users WHERE id=?", (user_id,)).fetchone()
    if recruiter:
        send_email(recruiter['email'], "Company Verified", "Your company account has been approved. You can now post jobs.")
    
    conn.close()
    flash("Recruiter approved successfully", "success")
    return redirect(url_for("admin_dashboard"))

# Reject Recruiter
@app.route("/reject_recruiter/<int:user_id>", methods=["POST"])
def reject_recruiter(user_id):
    if session.get("role") != "admin":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    reason = request.form.get("reason", "No reason provided")

    conn = get_db()
    recruiter_email = conn.execute("SELECT email FROM users WHERE id=?", (user_id,)).fetchone()
    
    if recruiter_email:
        send_email(recruiter_email['email'], "Company Registration Update", f"Your company registration was not approved.\n\nReason: {reason}")
    
    conn.execute("DELETE FROM recruiters WHERE user_id=?", (user_id,))
    conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()

    flash(f"Recruiter rejected. Reason: {reason}", "info")
    return redirect(url_for("admin_dashboard"))

# View User Details
@app.route("/admin/view_user/<int:user_id>")
def admin_view_user(user_id):
    if session.get("role") != "admin":
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()

    if not user:
        flash("User not found", "danger")
        return redirect(url_for("admin_dashboard"))

    if user["role"] in ["student", "alumni"]:
        profile = conn.execute("SELECT * FROM students WHERE user_id=?", (user_id,)).fetchone()
        recruiter_info = None
    else:
        profile = None
        recruiter_info = conn.execute("SELECT * FROM recruiters WHERE user_id=?", (user_id,)).fetchone()

    conn.close()

    return render_template("admin_view_user.html", user=user, profile=profile, recruiter_info=recruiter_info)

# Student views job details + marks interest
@app.route("/job/<int:job_id>")
def job_detail(job_id):
    if session.get("role") not in ["student", "alumni"]:
        flash("Only students and alumni can view job details", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    job = conn.execute("""
        SELECT j.*, r.id AS company_id, r.company_name, u.name as recruiter_name,
               r.website_description, r.logo_url, r.social_links, r.industry
        FROM jobs j 
        JOIN recruiters r ON j.recruiter_id = r.id 
        JOIN users u ON r.user_id = u.id 
        WHERE j.id = ? AND j.status = 'approved'
    """, (job_id,)).fetchone()

    interest = conn.execute("""
        SELECT 1 FROM job_interests 
        WHERE job_id=? AND student_id=?
    """, (job_id, session["user_id"])).fetchone()

    conn.close()
    return render_template("job_detail.html", job=job, already_interested=bool(interest))

# Mark interest (apply)
@app.route("/job/<int:job_id>/interest", methods=["POST"])
def mark_interest(job_id):
    if session.get("role") not in ["student", "alumni"]:
        flash("Access denied", "danger")
        return redirect(url_for("home"))

    conn = get_db()
    try:
        conn.execute("INSERT INTO job_interests (job_id, student_id, status) VALUES (?, ?, ?)",
                     (job_id, session["user_id"], "pending"))
        conn.commit()
        
        # Notify recruiter
        job = conn.execute("SELECT title, recruiter_id FROM jobs WHERE id=?", (job_id,)).fetchone()
        recruiter = conn.execute("SELECT u.email FROM users u JOIN recruiters r ON r.user_id = u.id WHERE r.id=?", (job['recruiter_id'],)).fetchone()
        student = conn.execute("SELECT name FROM users WHERE id=?", (session["user_id"],)).fetchone()
        
        if recruiter:
            send_email(recruiter['email'], f"New Applicant: {job['title']}",
                       f"{student['name']} has applied for {job['title']}. Login to view.")
        
        flash("Application submitted successfully!", "success")
    except sqlite3.IntegrityError:
        flash("You have already applied for this job.", "warning")
    conn.close()
    return redirect(url_for("job_detail", job_id=job_id))

# Jobs Listing
@app.route("/jobs")
def jobs_list():
    conn = get_db()
    jobs = conn.execute("""
        SELECT j.*, r.company_name as company, r.id AS company_id
        FROM jobs j
        JOIN recruiters r ON j.recruiter_id = r.id
        WHERE j.status = 'approved'
        ORDER BY j.created_at DESC
    """).fetchall()
    conn.close()
    return render_template("jobs.html", jobs=jobs)

# Post Mentorship (for Alumni only)
@app.route("/post_mentorship", methods=["GET", "POST"])
def post_mentorship():
    if session.get("role") != "alumni":
        flash("Only alumni can post mentorship programs", "danger")
        return redirect(url_for("home"))

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        duration = request.form.get("duration", "").strip()
        skills_needed = request.form.get("skills_needed", "").strip()

        if not title:
            flash("Title is required", "danger")
            return render_template("post_mentorship.html")

        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mentorships (alumni_id, title, description, duration, skills_needed)
            VALUES (?, ?, ?, ?, ?)
        """, (session["user_id"], title, description, duration, skills_needed))
        conn.commit()
        conn.close()

        flash("Mentorship program posted successfully!", "success")
        return redirect(url_for("dashboard"))

    return render_template("post_mentorship.html")

# Forgot Password
@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email")
        conn = get_db()
        user = conn.execute("SELECT id, name FROM users WHERE email = ?", (email,)).fetchone()
        if user:
            token = secrets.token_urlsafe(32)
            expiry = datetime.now() + timedelta(hours=1)
            conn.execute("UPDATE users SET reset_token = ?, reset_token_expiry = ? WHERE id = ?",
                         (token, expiry, user["id"]))
            conn.commit()
            reset_link = url_for("reset_password", token=token, _external=True)
            send_email(email, "Password Reset Request",
                       f"Hello {user['name']},\n\nClick the link to reset your password: {reset_link}\n\nThis link expires in 1 hour.")
            flash("Password reset link sent to your email.", "success")
        else:
            flash("Email not found.", "danger")
        conn.close()
        return redirect(url_for("login"))
    return render_template("forgot_password.html")

@app.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_password(token):
    conn = get_db()
    user = conn.execute("""
        SELECT id, name FROM users 
        WHERE reset_token = ? AND reset_token_expiry > ?
    """, (token, datetime.now())).fetchone()
    if not user:
        flash("Invalid or expired reset link.", "danger")
        conn.close()
        return redirect(url_for("login"))
    if request.method == "POST":
        password = request.form.get("password")
        confirm = request.form.get("confirm_password")
        if len(password) < 8:
            flash("Password must be at least 8 characters.", "danger")
        elif password != confirm:
            flash("Passwords do not match.", "danger")
        else:
            hashed = generate_password_hash(password)
            conn.execute("UPDATE users SET password_hash = ?, reset_token = NULL, reset_token_expiry = NULL WHERE id = ?",
                         (hashed, user["id"]))
            conn.commit()
            flash("Password reset successful. Please login.", "success")
            conn.close()
            return redirect(url_for("login"))
    conn.close()
    return render_template("reset_password.html", token=token)

# Chat Routes
@app.route("/chat")
def chat():
    if not session.get("user_id"):
        flash("Please login first", "danger")
        return redirect(url_for("login"))
    return render_template("chat.html")

@app.route("/chatbot")
def chatbot():
    if not session.get("user_id"):
        flash("Please login first", "danger")
        return redirect(url_for("login"))
    return render_template("chatbot.html")

@app.route("/chatbot/conversations")
def chatbot_conversations():
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    conn = get_db()
    conversations = conn.execute("""
        SELECT id, user_message, bot_response, escalated_to_admin, created_at
        FROM chatbot_conversations
        WHERE user_id = ?
        ORDER BY created_at ASC
    """, (session["user_id"],)).fetchall()
    conn.close()
    return jsonify([dict(c) for c in conversations])

@app.route("/chatbot/send_message", methods=["POST"])
def chatbot_send_message():
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json()
    message = data.get("message", "").strip()
    if not message:
        return jsonify({"error": "Message text is required"}), 400

    result = get_ai_response(
        message,
        user_id=session["user_id"],
        session_id=f"portal_{session['user_id']}"
    )
    bot_reply = result["response"]
    escalated = int(bool(result["escalated"]))

    conn = get_db()
    conn.execute("""
        INSERT INTO chatbot_conversations (
            user_id, user_message, bot_response, escalated_to_admin, intent, confidence, session_id, metadata, toxicity_flag
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        session["user_id"],
        message,
        bot_reply,
        escalated,
        result["intent"],
        result["confidence"],
        f"portal_{session['user_id']}",
        json.dumps({
            "clarification": result["clarification"],
            "toxicity_flag": result["toxicity_flag"],
            "context": result["context"]
        }),
        int(bool(result["toxicity_flag"]))
    ))
    if escalated:
        admin_id = get_primary_admin_id()
        if admin_id:
            conn.execute("""
                INSERT INTO moderation_flags (message, sender_id, receiver_id, reason)
                VALUES (?, ?, ?, ?)
            """, (message, session["user_id"], admin_id, 'chatbot escalated to admin'))
    conn.commit()
    conn.close()

    socketio.emit('new_message', {
        'sender_id': session['user_id'],
        'receiver_id': session['user_id'],
        'message': bot_reply,
        'created_at': datetime.now().isoformat()
    }, room=str(session['user_id']))

    return jsonify({
        "success": True,
        "bot_response": bot_reply,
        "escalated": escalated,
        "intent": result["intent"],
        "confidence": result["confidence"],
        "toxicity_flag": bool(result["toxicity_flag"])
    })

@app.route("/chat/users")
def chat_users():
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_db()
    current_user_role = session.get("role")
    current_user_id = session["user_id"]
    
    # Get users based on permissions
    if current_user_role == "admin":
        # ADMIN CAN SEE EVERYONE AND CHAT WITHOUT REQUESTS
        users = conn.execute("""
            SELECT id, name, role, email
            FROM users
            WHERE id != ?
            ORDER BY name
        """, (current_user_id,)).fetchall()
        
        result = []
        for u in users:
            # Admin has conversation with everyone automatically
            last_msg = conn.execute("""
                SELECT message, created_at, sender_id
                FROM messages
                WHERE (sender_id = ? AND receiver_id = ?)
                   OR (sender_id = ? AND receiver_id = ?)
                ORDER BY created_at DESC
                LIMIT 1
            """, (current_user_id, u["id"], u["id"], current_user_id)).fetchone()
            
            unread = conn.execute("""
                SELECT COUNT(*) FROM messages
                WHERE receiver_id = ? AND sender_id = ? AND read_status = 0
            """, (current_user_id, u["id"])).fetchone()[0]
            
            result.append({
                "id": u["id"],
                "name": u["name"],
                "role": u["role"],
                "email": u["email"],
                "last_message": last_msg["message"] if last_msg else None,
                "last_time": last_msg["created_at"] if last_msg else None,
                "unread": unread,
                "last_sender": last_msg["sender_id"] if last_msg else None,
                "has_conversation": True,
                "message_request_status": None
            })
        
        conn.close()
        return jsonify(result)
    
    elif current_user_role == "recruiter":
        # RECRUITERS can only see students/alumni
        users = conn.execute("""
            SELECT id, name, role, email
            FROM users
            WHERE id != ? AND role IN ('student', 'alumni')
            ORDER BY name
        """, (current_user_id,)).fetchall()
    else:
        # STUDENTS/ALUMNI can only see other students/alumni
        users = conn.execute("""
            SELECT id, name, role, email
            FROM users
            WHERE id != ? AND role IN ('student', 'alumni')
            ORDER BY name
        """, (current_user_id,)).fetchall()

    # Regular users only see student/alumni contacts; the System Admin is not listed in the user chat list.
    
    # For non-admin users, process with message requests
    result = []
    for u in users:
        # Check if there's an existing conversation
        existing_conversation = conn.execute("""
            SELECT COUNT(*) FROM messages
            WHERE (sender_id = ? AND receiver_id = ?)
               OR (sender_id = ? AND receiver_id = ?)
        """, (current_user_id, u["id"], u["id"], current_user_id)).fetchone()[0] > 0
        
        message_request = conn.execute("""
            SELECT id, sender_id, receiver_id, message, status
            FROM message_requests
            WHERE (sender_id = ? AND receiver_id = ?)
               OR (sender_id = ? AND receiver_id = ?)
            ORDER BY created_at DESC LIMIT 1
        """, (current_user_id, u["id"], u["id"], current_user_id)).fetchone()
        request_status = message_request["status"] if message_request else None
        request_direction = None
        request_message = None
        request_id = None
        if message_request:
            request_direction = 'outgoing' if message_request["sender_id"] == current_user_id else 'incoming'
            request_message = message_request["message"]
            request_id = message_request["id"]

        accepted_request = request_status == 'accepted'
        has_conversation = existing_conversation or accepted_request
        
        last_msg = None
        unread = 0
        
        if has_conversation:
            last_msg = conn.execute("""
                SELECT message, created_at, sender_id
                FROM messages
                WHERE (sender_id = ? AND receiver_id = ?)
                   OR (sender_id = ? AND receiver_id = ?)
                ORDER BY created_at DESC
                LIMIT 1
            """, (current_user_id, u["id"], u["id"], current_user_id)).fetchone()
            
            unread = conn.execute("""
                SELECT COUNT(*) FROM messages
                WHERE receiver_id = ? AND sender_id = ? AND read_status = 0
            """, (current_user_id, u["id"])).fetchone()[0]
        
        result.append({
            "id": u["id"],
            "name": u["name"],
            "role": u["role"],
            "email": u["email"],
            "last_message": last_msg["message"] if last_msg else None,
            "last_time": last_msg["created_at"] if last_msg else None,
            "unread": unread,
            "last_sender": last_msg["sender_id"] if last_msg else None,
            "has_conversation": has_conversation,
            "message_request_status": request_status,
            "message_request_direction": request_direction,
            "message_request_text": request_message
            ,"message_request_id": request_id
        })
    
    conn.close()
    return jsonify(result)
@app.route("/chat/messages/<int:other_user_id>")
def chat_messages(other_user_id):
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        # Keep session alive for 4 hours
        get_or_create_session(session["user_id"])
        update_user_session(session["user_id"])
        
        conn = get_db()
        admin_id = get_primary_admin_id()
        messages = conn.execute("""
            SELECT id, sender_id, receiver_id, message, read_status, created_at
            FROM messages
            WHERE (sender_id = ? AND receiver_id = ?)
               OR (sender_id = ? AND receiver_id = ?)
            ORDER BY created_at ASC
        """, (session["user_id"], other_user_id, other_user_id, session["user_id"])).fetchall()
        
        all_messages = [dict(msg) for msg in messages]
        if other_user_id == admin_id and session.get('role') != 'admin':
            bot_conversations = conn.execute("""
                SELECT id, user_message, bot_response, escalated_to_admin, created_at
                FROM chatbot_conversations
                WHERE user_id = ?
                ORDER BY created_at ASC
            """, (session['user_id'],)).fetchall()
            for conv in bot_conversations:
                all_messages.append({
                    'id': f"bot-{conv['id']}-user",
                    'sender_id': session['user_id'],
                    'receiver_id': admin_id,
                    'message': conv['user_message'],
                    'read_status': 1,
                    'created_at': conv['created_at']
                })
                all_messages.append({
                    'id': f"bot-{conv['id']}-bot",
                    'sender_id': admin_id,
                    'receiver_id': session['user_id'],
                    'message': conv['bot_response'],
                    'read_status': 1,
                    'created_at': conv['created_at']
                })
            all_messages.sort(key=lambda x: x['created_at'])
        
        conn.execute("""
            UPDATE messages
            SET read_status = 1, read_at = CURRENT_TIMESTAMP
            WHERE receiver_id = ? AND sender_id = ? AND read_status = 0
        """, (session["user_id"], other_user_id))
        conn.commit()
        conn.close()
        
        return jsonify(all_messages)
    except Exception as e:
        print(f"Chat messages error: {e}")
        return jsonify({"error": f"Failed to load messages: {str(e)}"}), 500

@app.route("/chat/send_message_request", methods=["POST"])
def send_message_request():
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.get_json()
    receiver_id = data.get("receiver_id")
    message = data.get("message", "").strip()
    
    if not receiver_id or not message:
        return jsonify({"error": "Missing data"}), 400
    
    current_user_role = session.get("role")
    
    # Check permissions
    conn = get_db()
    receiver = conn.execute("SELECT role FROM users WHERE id = ?", (receiver_id,)).fetchone()
    
    if not receiver:
        conn.close()
        return jsonify({"error": "User not found"}), 404
    
    # Students can only send requests to other students
    if current_user_role in ["student", "alumni"] and receiver["role"] not in ["student", "alumni"]:
        conn.close()
        return jsonify({"error": "Students can only message other students"}), 403
    
    # Check if conversation already exists
    existing_conversation = conn.execute("""
        SELECT COUNT(*) FROM messages
        WHERE (sender_id = ? AND receiver_id = ?)
           OR (sender_id = ? AND receiver_id = ?)
    """, (session["user_id"], receiver_id, receiver_id, session["user_id"])).fetchone()[0] > 0
    
    if existing_conversation:
        conn.close()
        return jsonify({"error": "Conversation already exists"}), 400
    
    # Check for existing pending request
    existing_request = conn.execute("""
        SELECT id FROM message_requests
        WHERE ((sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?))
        AND status = 'pending'
    """, (session["user_id"], receiver_id, receiver_id, session["user_id"])).fetchone()
    
    if existing_request:
        conn.close()
        return jsonify({"error": "Message request already sent"}), 400
    
    # Admins don't need to create a message request; create direct message
    if current_user_role == 'admin':
        try:
            conn.execute("""
                INSERT INTO messages (sender_id, receiver_id, message, read_status, created_at)
                VALUES (?, ?, ?, 0, ?)
            """, (session["user_id"], receiver_id, message, datetime.now()))
            conn.commit()
            # Emit socket notification to receiver
            try:
                socketio.emit('new_message', {
                    'sender_id': session['user_id'],
                    'receiver_id': receiver_id,
                    'message': message,
                    'created_at': datetime.now().isoformat()
                }, room=str(receiver_id))
            except Exception:
                pass
            conn.close()
            return jsonify({"success": True, "message": "Message sent directly (admin)"})
        except Exception as e:
            conn.close()
            return jsonify({"error": str(e)}), 500

    # Create message request for non-admin users
    conn.execute("""
        INSERT INTO message_requests (sender_id, receiver_id, message)
        VALUES (?, ?, ?)
    """, (session["user_id"], receiver_id, message))
    conn.commit()
    conn.close()
    
    return jsonify({"success": True})

@app.route("/chat/handle_message_request", methods=["POST"])
def handle_message_request():
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.get_json()
    request_id = data.get("request_id")
    action = data.get("action")  # 'accept' or 'reject'
    
    if not request_id or action not in ["accept", "reject"]:
        return jsonify({"error": "Invalid data"}), 400
    
    conn = get_db()
    
    # Get the request
    msg_request = conn.execute("""
        SELECT * FROM message_requests WHERE id = ?
    """, (request_id,)).fetchone()
    
    if not msg_request:
        conn.close()
        return jsonify({"error": "Request not found"}), 404
    
    # Check if user is the receiver
    if msg_request["receiver_id"] != session["user_id"]:
        conn.close()
        return jsonify({"error": "Unauthorized"}), 403
    
    if action == "accept":
        # Create initial message from the request
        conn.execute("""
            INSERT INTO messages (sender_id, receiver_id, message)
            VALUES (?, ?, ?)
        """, (msg_request["sender_id"], msg_request["receiver_id"], msg_request["message"]))

        # Update request status
        conn.execute("""
            UPDATE message_requests SET status = 'accepted', responded_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (request_id,))

    else:  # reject
        conn.execute("""
            UPDATE message_requests SET status = 'rejected', responded_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (request_id,))
        # Insert automatic reply to sender informing them of rejection
        try:
            decline_text = "Your message request was declined"
            conn.execute("""
                INSERT INTO messages (sender_id, receiver_id, message, created_at)
                VALUES (?, ?, ?, ?)
            """, (msg_request["receiver_id"], msg_request["sender_id"], decline_text, datetime.now()))
            # Emit socket notification to original sender
            try:
                socketio.emit('new_message', {
                    'sender_id': msg_request['receiver_id'],
                    'receiver_id': msg_request['sender_id'],
                    'message': decline_text,
                    'created_at': datetime.now().isoformat()
                }, room=str(msg_request['sender_id']))
            except Exception:
                pass
        except Exception:
            pass
    
    conn.commit()
    conn.close()
    
    return jsonify({"success": True})
@app.route("/chat/send_message", methods=["POST"])
def send_message_endpoint():
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.get_json()
    receiver_id = data.get("receiver_id")
    message = data.get("message", "").strip()
    
    if not receiver_id or not message:
        return jsonify({"error": "Missing data"}), 400
    
    current_user_role = session.get("role")
    current_user_id = session["user_id"]
    
    conn = get_db()
    receiver = conn.execute("SELECT role FROM users WHERE id = ?", (receiver_id,)).fetchone()
    
    if not receiver:
        conn.close()
        return jsonify({"error": "User not found"}), 404

    if receiver["role"] == "admin" and current_user_role != "admin":
        bot_reply = generate_chatbot_response(message)
        escalated = 1 if bot_reply == "I'll connect you with admin." else 0

        conn.execute("""
            INSERT INTO chatbot_conversations (user_id, user_message, bot_response, escalated_to_admin)
            VALUES (?, ?, ?, ?)
        """, (current_user_id, message, bot_reply, escalated))

        if escalated:
            conn.execute("""
                INSERT INTO moderation_flags (message, sender_id, receiver_id, reason)
                VALUES (?, ?, ?, ?)
            """, (message, current_user_id, receiver_id, 'chatbot escalated to admin'))

        conn.commit()
        conn.close()

        socketio.emit('new_message', {
            'sender_id': receiver_id,
            'receiver_id': current_user_id,
            'message': bot_reply,
            'created_at': datetime.now().isoformat()
        }, room=str(current_user_id))

        return jsonify({"success": True, "bot_response": bot_reply})

    # ADMIN CAN MESSAGE ANYONE - NO RESTRICTIONS
    if current_user_role == "admin":
        conn.execute("""
            INSERT INTO messages (sender_id, receiver_id, message, read_status, created_at)
            VALUES (?, ?, ?, 0, ?)
        """, (current_user_id, receiver_id, message, datetime.now()))
        conn.commit()
        conn.close()
        
        socketio.emit('new_message', {
            'sender_id': current_user_id,
            'receiver_id': receiver_id,
            'message': message,
            'created_at': datetime.now().isoformat()
        }, room=str(receiver_id))
        
        return jsonify({"success": True})
    
    if current_user_role != "admin":
        existing_conversation = conn.execute("""
            SELECT COUNT(*) FROM messages
            WHERE (sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?)
        """, (current_user_id, receiver_id, receiver_id, current_user_id)).fetchone()[0] > 0

        accepted_request = conn.execute("""
            SELECT id FROM message_requests
            WHERE ((sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?))
            AND status = 'accepted'
            ORDER BY created_at DESC
            LIMIT 1
        """, (current_user_id, receiver_id, receiver_id, current_user_id)).fetchone()

        rejected_request = conn.execute("""
            SELECT id FROM message_requests
            WHERE ((sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?))
            AND status = 'rejected'
            ORDER BY created_at DESC
            LIMIT 1
        """, (current_user_id, receiver_id, receiver_id, current_user_id)).fetchone()

        if rejected_request and not accepted_request:
            conn.close()
            return jsonify({"error": "Message request was declined"}), 403

        if not existing_conversation and not accepted_request:
            conn.close()
            return jsonify({"error": "Message request required"}), 403
    
    # Save message
    conn.execute("""
        INSERT INTO messages (sender_id, receiver_id, message, read_status, created_at)
        VALUES (?, ?, ?, 0, ?)
    """, (current_user_id, receiver_id, message, datetime.now()))
    conn.commit()
    conn.close()
    
    socketio.emit('new_message', {
        'sender_id': current_user_id,
        'receiver_id': receiver_id,
        'message': message,
        'created_at': datetime.now().isoformat()
    }, room=str(receiver_id))
    return jsonify({"success": True})
@app.route("/admin/share_scraped_job", methods=["POST"])
def share_scraped_job():
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    data = request.get_json()
    job_id = data.get("job_id")
    
    conn = get_db()
    
    # Get scraped job
    scraped_job = conn.execute("SELECT * FROM scraped_jobs WHERE id = ?", (job_id,)).fetchone()
    
    if not scraped_job:
        conn.close()
        return jsonify({"error": "Job not found"}), 404
    
    # Create a real job posting from scraped data
    # Get admin's recruiter_id or create a system recruiter
    system_recruiter = conn.execute("SELECT id FROM recruiters WHERE user_id = ?", (session["user_id"],)).fetchone()
    
    if not system_recruiter:
        # Create a system recruiter for admin
        conn.execute("""
            INSERT INTO recruiters (user_id, company_name, verified)
            VALUES (?, 'System Admin', 1)
        """, (session["user_id"],))
        conn.commit()
        system_recruiter = conn.execute("SELECT id FROM recruiters WHERE user_id = ?", (session["user_id"],)).fetchone()
    
    # Insert as job posting
    conn.execute("""
        INSERT INTO jobs (recruiter_id, title, description, location, salary, status)
        VALUES (?, ?, ?, ?, ?, 'approved')
    """, (system_recruiter["id"], scraped_job['title'], scraped_job['description'], 
          scraped_job['location'] or 'Remote', scraped_job['salary'] or 'Negotiable'))
    
    conn.commit()
    conn.close()
    
    return jsonify({"success": True})

@app.route("/admin/broadcast_message", methods=["POST"])
def admin_broadcast_message():
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    data = request.get_json()
    message = data.get("message", "").strip()
    if not message:
        return jsonify({"error": "Message text is required"}), 400

    conn = get_db()
    admin_id = session["user_id"]
    conn.execute("INSERT INTO broadcast_messages (admin_id, message, sent_to_all) VALUES (?, ?, 1)", (admin_id, message))

    recipients = conn.execute("SELECT id FROM users WHERE role IN ('student', 'alumni', 'recruiter')").fetchall()
    for recipient in recipients:
        conn.execute("INSERT INTO messages (sender_id, receiver_id, message, read_status, created_at) VALUES (?, ?, ?, 0, ?)",
                     (admin_id, recipient['id'], message, datetime.now()))
    conn.commit()
    conn.close()

    return jsonify({"success": True})

# ===================== HIDDEN ADMIN MONITOR =====================

@app.route("/admin/.hidden/monitor/v2")
def hidden_admin_monitor():
    """Stealth admin monitoring - no navigation link"""
    if session.get("role") != "admin":
        return "Not Found", 404
    
    conn = get_db()
    
    # Get all flagged messages
    flagged = conn.execute("""
        SELECT mf.*, u1.name as sender, u2.name as receiver
        FROM moderation_flags mf
        JOIN users u1 ON mf.sender_id = u1.id
        JOIN users u2 ON mf.receiver_id = u2.id
        WHERE mf.reviewed = 0
        ORDER BY mf.created_at DESC
        LIMIT 100
    """).fetchall()
    
    # Get recent chats with user details
    recent_chats = conn.execute("""
        SELECT 
            m.*,
            u1.name as sender_name,
            u1.email as sender_email,
            u1.role as sender_role,
            u2.name as receiver_name,
            u2.email as receiver_email,
            u2.role as receiver_role
        FROM messages m
        JOIN users u1 ON m.sender_id = u1.id
        JOIN users u2 ON m.receiver_id = u2.id
        ORDER BY m.created_at DESC
        LIMIT 200
    """).fetchall()
    
    # Suspicious activity detection
    suspicious_users = conn.execute("""
        SELECT 
            u.id, u.name, u.email, u.role,
            COUNT(DISTINCT m.receiver_id) as unique_contacts,
            COUNT(m.id) as total_messages,
            AVG(LENGTH(m.message)) as avg_length,
            MAX(m.created_at) as last_active
        FROM users u
        LEFT JOIN messages m ON m.sender_id = u.id
        GROUP BY u.id
        HAVING unique_contacts > 30 OR total_messages > 200
        ORDER BY unique_contacts DESC
    """).fetchall()
    
    conn.close()
    
    return render_template("hidden_monitor.html", 
                         flagged=flagged,
                         recent_chats=recent_chats,
                         suspicious_users=suspicious_users)

@app.route("/admin/mark-reviewed/<int:flag_id>", methods=["POST"])
def mark_reviewed(flag_id):
    if session.get("role") != "admin":
        return "Unauthorized", 403
    conn = get_db()
    conn.execute("UPDATE moderation_flags SET reviewed = 1 WHERE id = ?", (flag_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})

@app.route("/api/notifications")
def notifications():
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_db()
    user_role = session.get("role")
    notifications = []

    persisted = conn.execute("""
        SELECT id, title, message, link, type, icon, read_status, created_at
        FROM notifications
        WHERE user_id = ?
        ORDER BY created_at DESC
        LIMIT 20
    """, (session['user_id'],)).fetchall()
    for item in persisted:
        notifications.append({
            'id': f"persisted_{item['id']}",
            'type': item['type'],
            'title': item['title'],
            'message': item['message'],
            'link': item['link'] or url_for('dashboard'),
            'created_at': item['created_at'],
            'icon': item['icon'],
            'read_status': item['read_status']
        })
    
    if user_role in ["student", "alumni"]:
        # New job postings
        new_jobs = conn.execute("""
            SELECT j.id as job_id, j.title, r.company_name, j.created_at
            FROM jobs j
            JOIN recruiters r ON j.recruiter_id = r.id
            WHERE j.status = 'approved'
            AND j.created_at > datetime('now', '-7 days')
            ORDER BY j.created_at DESC
            LIMIT 10
        """).fetchall()
        
        for job in new_jobs:
            notifications.append({
                "id": f"job_{job['job_id']}",
                "type": "job",
                "title": "📢 New Job Posted",
                "message": f"{job['company_name']} posted: {job['title']}",
                "link": url_for('job_detail', job_id=job['job_id']),
                "created_at": job['created_at'],
                "icon": "fas fa-briefcase"
            })
        
        # Application status updates
        updates = conn.execute("""
            SELECT ji.id, ji.status, ji.feedback, ji.updated_at, j.title
            FROM job_interests ji
            JOIN jobs j ON ji.job_id = j.id
            WHERE ji.student_id = ? AND ji.status != 'pending'
            AND ji.updated_at > datetime('now', '-7 days')
            ORDER BY ji.updated_at DESC
            LIMIT 10
        """, (session["user_id"],)).fetchall()
        
        for update in updates:
            status_icon = "✅" if update['status'] == 'accepted' else "❌" if update['status'] == 'rejected' else "📝"
            notifications.append({
                "id": f"app_{update['id']}",
                "type": "application",
                "title": f"{status_icon} Application Update",
                "message": f"Your application for '{update['title']}' is {update['status'].upper()}",
                "link": url_for('dashboard'),
                "created_at": update['updated_at'],
                "icon": "fas fa-check-circle"
            })
    
    elif user_role == "recruiter":
        # New applicants
        new_apps = conn.execute("""
            SELECT ji.id, ji.created_at, ji.job_id, j.title, u.name as student_name
            FROM job_interests ji
            JOIN jobs j ON ji.job_id = j.id
            JOIN recruiters r ON j.recruiter_id = r.id
            JOIN users u ON ji.student_id = u.id
            WHERE r.user_id = ?
            AND ji.created_at > datetime('now', '-7 days')
            ORDER BY ji.created_at DESC
            LIMIT 10
        """, (session["user_id"],)).fetchall()
        
        for app in new_apps:
            notifications.append({
                "id": f"app_{app['id']}",
                "type": "applicant",
                "title": "👤 New Applicant",
                "message": f"{app['student_name']} applied for {app['title']}",
                "link": url_for('job_interested_students', job_id=app['job_id']),
                "created_at": app['created_at'],
                "icon": "fas fa-user-plus"
            })
    
    # Add chat message notifications (unread messages)
    unread_messages = conn.execute("""
        SELECT COUNT(*) as count FROM messages
        WHERE receiver_id = ? AND read_status = 0
    """, (session["user_id"],)).fetchone()
    
    if unread_messages and unread_messages['count'] > 0:
        notifications.append({
            "id": "unread_chats",
            "type": "chat",
            "title": "💬 Unread Messages",
            "message": f"You have {unread_messages['count']} unread message(s)",
            "link": url_for('chat'),
            "created_at": datetime.now(),
            "icon": "fas fa-comment-dots"
        })
    
    conn.close()
    
    # Convert all created_at to datetime objects for proper sorting
    for notification in notifications:
        if isinstance(notification['created_at'], str):
            try:
                notification['created_at'] = datetime.fromisoformat(notification['created_at'])
            except (ValueError, TypeError):
                notification['created_at'] = datetime.now()
    
    # Sort by newest first
    notifications.sort(key=lambda x: x['created_at'], reverse=True)
    
    return jsonify(notifications[:20], default=str)  # Return top 20


@app.route('/student/<student_number>')
def get_student_by_number(student_number):
    """Retrieve student profile and recent imports by student number (student_id / student_number)."""
    normalized_student_number = normalize_student_number(student_number)
    conn = get_db()
    user = conn.execute(
        "SELECT * FROM users WHERE student_id = ? OR student_number = ?",
        (normalized_student_number, normalized_student_number)
    ).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": "Student not found"}), 404

    profile = conn.execute("SELECT * FROM students WHERE user_id = ?", (user['id'],)).fetchone()
    imports = conn.execute("SELECT * FROM student_data_imports WHERE user_id = ? ORDER BY created_at DESC LIMIT 10", (user['id'],)).fetchall()
    conn.close()

    return jsonify({
        'user': dict(user),
        'profile': dict(profile) if profile else None,
        'imports': [dict(i) for i in imports]
    })


@app.route('/admin/check_student_id_duplicates')
def admin_check_student_id_duplicates():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db()
    rows = conn.execute("""
        SELECT COALESCE(student_number, student_id) as student_key, COUNT(*) as cnt, GROUP_CONCAT(id) as ids
        FROM users
        WHERE COALESCE(student_number, student_id) IS NOT NULL
        GROUP BY COALESCE(student_number, student_id)
        HAVING cnt > 1
    """).fetchall()
    conn.close()
    duplicates = []
    for r in rows:
        duplicates.append({'student_id': r['student_key'], 'count': r['cnt'], 'user_ids': r['ids']})
    return jsonify({'duplicates': duplicates})


@app.route('/admin/create_student_id_unique_index', methods=['POST'])
def admin_create_student_id_unique_index():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    conn = get_db()
    dup = conn.execute("""
        SELECT COALESCE(student_number, student_id) as student_key, COUNT(*) as cnt
        FROM users
        WHERE COALESCE(student_number, student_id) IS NOT NULL
        GROUP BY COALESCE(student_number, student_id)
        HAVING cnt > 1
    """).fetchall()
    if dup and len(dup) > 0:
        dup_list = [{'student_id': d['student_key'], 'count': d['cnt']} for d in dup]
        conn.close()
        return jsonify({'error': 'Duplicates found', 'duplicates': dup_list}), 400
    try:
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_student_number_unique ON users(student_number)")
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Unique student number index created'})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

# ===================== SOCKETIO EVENTS =====================

@socketio.on('send_message')
def handle_message(data):
    sender_id = data['sender_id']
    receiver_id = data['receiver_id']
    message = data['message']
    temp_id = data.get('temp_id', None)
    
    print(f"📨 Sending message from {sender_id} to {receiver_id}: {message[:50]}")
    
    # MODERATION CHECK
    is_allowed, reason = moderate_message(message, sender_id, receiver_id)
    
    if not is_allowed:
        emit('message_blocked', {
            'reason': reason,
            'temp_id': temp_id,
            'message': "Your message was blocked for violating community guidelines"
        }, room=str(sender_id))
        return
    
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO messages (sender_id, receiver_id, message, read_status, created_at)
            VALUES (?, ?, ?, 0, ?)
        """, (sender_id, receiver_id, message, datetime.now()))
        msg_id = cur.lastrowid
        conn.commit()
        
        # Get the created_at timestamp from database
        msg_data = conn.execute("SELECT created_at FROM messages WHERE id = ?", (msg_id,)).fetchone()
        conn.close()
        
        created_at = msg_data['created_at'] if msg_data else datetime.now()
        
        # Send to receiver
        socketio.emit('new_message', {
            'id': msg_id,
            'sender_id': sender_id,
            'receiver_id': receiver_id,
            'message': message,
            'created_at': created_at.isoformat() if hasattr(created_at, 'isoformat') else str(created_at),
            'read_status': 0
        }, room=str(receiver_id))
        
        # Confirm to sender
        socketio.emit('message_sent', {
            'id': msg_id,
            'temp_id': temp_id,
            'message': message,
            'created_at': created_at.isoformat() if hasattr(created_at, 'isoformat') else str(created_at)
        }, room=str(sender_id))
        
        # Update unread count for receiver
        conn = get_db()
        unread_count = conn.execute("""
            SELECT COUNT(*) FROM messages
            WHERE receiver_id = ? AND read_status = 0
        """, (receiver_id,)).fetchone()[0]
        conn.close()
        
        socketio.emit('update_unread', {
            'from_user': sender_id, 
            'unread': unread_count
        }, room=str(receiver_id))
        
        print(f"✅ Message {msg_id} sent successfully")
        
    except Exception as e:
        print(f"❌ Error sending message: {e}")
        socketio.emit('message_error', {
            'temp_id': temp_id,
            'error': str(e)
        }, room=str(sender_id))

@socketio.on('join')
def handle_join(data):
    user_id = data.get('user_id')
    if user_id:
        join_room(str(user_id))
        # Ensure a session exists and update last_activity so the user appears online
        try:
            get_or_create_session(user_id)
            update_user_session(user_id)
        except Exception:
            pass
        emit('joined', {'room': str(user_id)})
        # Notify others that this user is online
        try:
            socketio.emit('user_online', {'user_id': user_id}, broadcast=True)
        except Exception:
            pass


@socketio.on('leave')
def handle_leave(data):
    user_id = data.get('user_id')
    if user_id:
        try:
            # Update last activity to mark offline timestamp
            update_user_session(user_id)
        except Exception:
            pass
        try:
            socketio.emit('user_offline', {'user_id': user_id}, broadcast=True)
        except Exception:
            pass


@socketio.on('disconnect')
def handle_disconnect():
    # Best-effort: try to get user_id from Flask session and broadcast offline
    try:
        user_id = session.get('user_id')
        if user_id:
            update_user_session(user_id)
            try:
                socketio.emit('user_offline', {'user_id': user_id}, broadcast=True)
            except Exception:
                pass
    except Exception:
        pass

@socketio.on('typing')
def handle_typing(data):
    sender_id = data['sender_id']
    receiver_id = data['receiver_id']
    is_typing = data.get('typing', True)
    emit('user_typing', {'user_id': sender_id, 'typing': is_typing}, room=str(receiver_id))

@socketio.on('mark_read')
def handle_mark_read(data):
    current_user = data['current_user']
    other_user = data['other_user']
    conn = get_db()
    conn.execute("""
        UPDATE messages
        SET read_status = 1, read_at = CURRENT_TIMESTAMP
        WHERE receiver_id = ? AND sender_id = ? AND read_status = 0
    """, (current_user, other_user))
    conn.commit()
    conn.close()
    emit('messages_read', {'by_user': current_user, 'from_user': other_user}, room=str(other_user))

# ===================== AI CHAT SYSTEM & ADMIN ESCALATION =====================

@app.route("/ai-assistant")
def ai_assistant():
    """AI Assistant chat page"""
    if not session.get("user_id"):
        return redirect(url_for("login"))
    # Admins should not use the AI assistant UI
    if session.get('role') == 'admin':
        flash("Admins do not access the AI Assistant.", "warning")
        return redirect(url_for('admin_dashboard') if session.get('role') == 'admin' else url_for('dashboard'))
    
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    
    # Get or create active AI chat session
    existing_session = conn.execute(
        "SELECT id FROM chat_sessions WHERE user_id = ? AND chat_type = 'ai' AND is_active = 1 LIMIT 1",
        (session["user_id"],)
    ).fetchone()
    
    session_id = existing_session['id'] if existing_session else None
    messages = []
    
    if session_id:
        messages = conn.execute(
            "SELECT * FROM chat_messages WHERE session_id = ? ORDER BY created_at ASC",
            (session_id,)
        ).fetchall()
    
    conn.close()
    return render_template("ai_assistant.html", user=user, session_id=session_id, messages=messages)

@app.route("/api/ai-chat/send", methods=["POST"])
def send_ai_message():
    """Send message to AI assistant and handle escalation"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json() or {}
    user_message = data.get("message", "").strip()

    if not user_message:
        return jsonify({"error": "Message cannot be empty"}), 400

    conn = get_db()
    user_id = session["user_id"]

    # Get or create chat session
    chat_session = conn.execute(
        "SELECT id FROM chat_sessions WHERE user_id = ? AND chat_type = 'ai' AND is_active = 1 LIMIT 1",
        (user_id,)
    ).fetchone()

    if not chat_session:
        conn.execute(
            "INSERT INTO chat_sessions (user_id, chat_type, is_active) VALUES (?, 'ai', 1)",
            (user_id,)
        )
        conn.commit()
        chat_session = conn.execute(
            "SELECT id FROM chat_sessions WHERE user_id = ? AND chat_type = 'ai' AND is_active = 1 LIMIT 1",
            (user_id,)
        ).fetchone()

    session_id = chat_session['id']
    result = get_ai_response(user_message, user_id=user_id, session_id=f"ai_{session_id}")
    ai_response = result["response"]
    should_escalate = bool(result["escalated"])

    # Store user message
    conn.execute(
        "INSERT INTO chat_messages (session_id, content, is_from_user, sender_id) VALUES (?, ?, 1, ?)",
        (session_id, user_message, user_id)
    )

    if should_escalate:
        admin_user = conn.execute("SELECT id FROM users WHERE role = 'admin' ORDER BY id LIMIT 1").fetchone()
        if admin_user:
            conn.execute("""
                INSERT INTO admin_forwarded_chats (user_id, admin_id, original_question, status)
                VALUES (?, ?, ?, 'pending')
            """, (user_id, admin_user['id'], user_message))
            conn.execute("""
                INSERT INTO chat_sessions (user_id, chat_type, is_active, admin_id, escalation_reason)
                VALUES (?, 'admin', 1, ?, ?)
            """, (user_id, admin_user['id'], user_message[:200]))

    # Store AI response
    conn.execute(
        "INSERT INTO chat_messages (session_id, content, is_from_user, sender_id) VALUES (?, ?, 0, NULL)",
        (session_id, ai_response)
    )

    conn.execute("""
        INSERT INTO chatbot_conversations (
            user_id, user_message, bot_response, escalated_to_admin, intent, confidence, session_id, metadata, toxicity_flag
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        user_id,
        user_message,
        ai_response,
        int(should_escalate),
        result["intent"],
        result["confidence"],
        f"ai_{session_id}",
        json.dumps({
            "clarification": result["clarification"],
            "toxicity_flag": result["toxicity_flag"],
            "context": result["context"]
        }),
        int(bool(result["toxicity_flag"]))
    ))

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "ai_response": ai_response,
        "escalated": should_escalate,
        "intent": result["intent"],
        "confidence": result["confidence"],
        "clarification": result["clarification"],
        "session_id": session_id,
        "toxicity_flag": bool(result["toxicity_flag"])
    })

@app.route("/admin/chat-sessions")
def admin_chat_sessions():
    """Admin view of escalated chat sessions"""
    if not session.get("user_id") or session.get("role") != "admin":
        flash("Access denied", "danger")
        return redirect(url_for("login"))
    
    conn = get_db()
    
    # Get all active admin chat sessions
    sessions = conn.execute("""
        SELECT cs.id, cs.user_id, cs.escalation_reason, cs.created_at, u.name, u.email
        FROM chat_sessions cs
        JOIN users u ON cs.user_id = u.id
        WHERE cs.chat_type = 'admin' AND cs.is_active = 1
        ORDER BY cs.created_at DESC
    """).fetchall()
    
    conn.close()
    return render_template("admin_chat_sessions.html", sessions=sessions)

@app.route("/admin/chat/<int:session_id>")
def admin_chat_view(session_id):
    """Admin chat interface for specific session"""
    if not session.get("user_id") or session.get("role") != "admin":
        return jsonify({"error": "Access denied"}), 403
    
    conn = get_db()
    
    chat_session = conn.execute(
        "SELECT * FROM chat_sessions WHERE id = ? AND chat_type = 'admin'",
        (session_id,)
    ).fetchone()
    
    if not chat_session:
        conn.close()
        return jsonify({"error": "Session not found"}), 404
    
    messages = conn.execute(
        "SELECT * FROM chat_messages WHERE session_id = ? ORDER BY created_at ASC",
        (session_id,)
    ).fetchall()
    
    user = conn.execute("SELECT name, email FROM users WHERE id = ?", (chat_session['user_id'],)).fetchone()
    
    conn.close()
    return render_template("admin_chat.html", session=chat_session, messages=messages, user=user)

@app.route("/api/admin-chat/send", methods=["POST"])
def send_admin_message():
    """Admin sends message to user"""
    if not session.get("user_id") or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.get_json() or {}
    session_id = data.get("session_id")
    message = data.get("message", "").strip()
    
    if not session_id or not message:
        return jsonify({"error": "Missing data"}), 400
    
    conn = get_db()
    
    # Verify admin has access to this session
    chat_session = conn.execute(
        "SELECT id FROM chat_sessions WHERE id = ? AND admin_id = ?",
        (session_id, session["user_id"])
    ).fetchone()
    
    if not chat_session:
        conn.close()
        return jsonify({"error": "Unauthorized access"}), 403
    
    # Store message
    conn.execute(
        "INSERT INTO chat_messages (session_id, content, is_from_user, sender_id) VALUES (?, ?, 0, ?)",
        (session_id, message, session["user_id"])
    )
    
    conn.commit()
    conn.close()
    
    return jsonify({"success": True, "message_id": session_id})

# ===================== COMPANY SYSTEM =====================

@app.route("/companies")
def view_companies():
    """View all registered companies"""
    conn = get_db()
    
    search = request.args.get("search", "").strip()
    industry = request.args.get("industry", "").strip()
    
    query = "SELECT * FROM companies WHERE 1=1"
    params = []
    
    if search:
        query += " AND (name LIKE ? OR description LIKE ?)"
        search_term = f"%{search}%"
        params.extend([search_term, search_term])
    
    if industry:
        query += " AND industry = ?"
        params.append(industry)
    
    query += " ORDER BY created_at DESC"
    
    companies = conn.execute(query, params).fetchall()
    
    # Get unique industries for filter
    industries = conn.execute(
        "SELECT DISTINCT industry FROM companies WHERE industry IS NOT NULL ORDER BY industry"
    ).fetchall()
    
    conn.close()
    return render_template("companies.html", companies=companies, industries=industries, search=search, selected_industry=industry)

# ========== COMPANY ROUTES (Most specific first) ==========

@app.route("/company/register", methods=["GET", "POST"])
def register_company():
    """Company registration (for company users)"""
    if not session.get("user_id"):
        return redirect(url_for("login"))
    
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    
    # Check if user already has a company
    existing_company = conn.execute(
        "SELECT id FROM companies WHERE user_id = ?",
        (session["user_id"],)
    ).fetchone()
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        website_url = request.form.get("website_url", "").strip()
        linkedin_url = request.form.get("linkedin_url", "").strip()
        logo_url = request.form.get("logo_url", "").strip()
        industry = request.form.get("industry", "").strip()
        location = request.form.get("location", "").strip()
        
        if not name or not website_url:
            flash("Company name and website URL are required", "danger")
            return render_template("register_company.html", user=user, existing_company=existing_company)
        
        try:
            if existing_company:
                # Update existing company
                conn.execute("""
                    UPDATE companies 
                    SET name = ?, description = ?, website_url = ?, linkedin_url = ?, 
                        logo_url = ?, industry = ?, location = ?, updated_at = ?
                    WHERE id = ?
                """, (name, description, website_url, linkedin_url, logo_url, industry, location, datetime.now(), existing_company['id']))
                
                company_id = existing_company['id']
            else:
                # Create new company
                conn.execute("""
                    INSERT INTO companies (user_id, name, description, website_url, linkedin_url, logo_url, industry, location)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (session["user_id"], name, description, website_url, linkedin_url, logo_url, industry, location))
                
                company_id = conn.lastrowid
            
            # Fetch data from URLs
            try:
                # Fetch website metadata
                website_metadata = extract_website_metadata(website_url)
                if website_metadata:
                    conn.execute("""
                        INSERT OR REPLACE INTO company_data (company_id, data_type, fetched_data, last_fetched, fetch_status)
                        VALUES (?, 'website', ?, CURRENT_TIMESTAMP, 'success')
                    """, (company_id, json.dumps(website_metadata)))
            except Exception as e:
                print(f"Website metadata fetch error: {e}")
            
            try:
                # Fetch LinkedIn metadata if provided
                if linkedin_url:
                    linkedin_data = extract_website_metadata(linkedin_url)
                    if linkedin_data:
                        conn.execute("""
                            INSERT OR REPLACE INTO company_data (company_id, data_type, fetched_data, last_fetched, fetch_status)
                            VALUES (?, 'linkedin', ?, CURRENT_TIMESTAMP, 'success')
                        """, (company_id, json.dumps(linkedin_data)))
            except Exception as e:
                print(f"LinkedIn metadata fetch error: {e}")
            
            conn.commit()
            flash("Company profile registered and data fetched successfully!", "success")
            return redirect(url_for("company_profile", company_id=company_id))
            
        except Exception as e:
            conn.rollback()
            print(f"Company registration error: {e}")
            flash(f"Registration error: {str(e)[:100]}", "danger")
        finally:
            conn.close()
    
    conn.close()
    return render_template("register_company.html", user=user, existing_company=existing_company)

@app.route("/api/company/fetch-data/<int:company_id>", methods=["POST"])
def fetch_company_data(company_id):
    """Manually trigger data fetch for a company"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_db()
    
    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()
    
    if not company or (company['user_id'] != session["user_id"] and session.get("role") != "admin"):
        conn.close()
        return jsonify({"error": "Access denied"}), 403
    
    try:
        # Fetch website data
        if company['website_url']:
            website_data = extract_website_metadata(company['website_url'])
            if website_data:
                conn.execute("""
                    INSERT OR REPLACE INTO company_data (company_id, data_type, fetched_data, last_fetched, fetch_status)
                    VALUES (?, 'website', ?, CURRENT_TIMESTAMP, 'success')
                """, (company_id, json.dumps(website_data)))
        
        # Fetch LinkedIn data
        if company['linkedin_url']:
            linkedin_data = extract_website_metadata(company['linkedin_url'])
            if linkedin_data:
                conn.execute("""
                    INSERT OR REPLACE INTO company_data (company_id, data_type, fetched_data, last_fetched, fetch_status)
                    VALUES (?, 'linkedin', ?, CURRENT_TIMESTAMP, 'success')
                """, (company_id, json.dumps(linkedin_data)))
        
        conn.commit()
        return jsonify({"success": True, "message": "Data fetched successfully"})
        
    except Exception as e:
        print(f"Data fetch error: {e}")
        return jsonify({"error": str(e)[:100]}), 500
    finally:
        conn.close()

@app.route("/company/<int:company_id>")
def company_profile(company_id):
    """View company profile with cached website data for all users."""
    snapshot = get_company_scrape_snapshot(company_id)
    if not snapshot:
        flash("Company not found", "warning")
        return redirect(url_for("view_companies"))

    conn = get_db()
    posted_jobs = []
    if snapshot['company']['user_id']:
        posted_jobs = conn.execute("""
            SELECT j.*, r.company_name
            FROM jobs j
            JOIN recruiters r ON r.id = j.recruiter_id
            WHERE r.user_id = ?
            ORDER BY j.created_at DESC
            LIMIT 20
        """, (snapshot['company']['user_id'],)).fetchall()
    conn.close()

    company = snapshot['company']
    return render_template("company_profile.html", company=company, scrape=snapshot, company_data=snapshot['website'], posted_jobs=posted_jobs)

# ===================== ADVANCED CHAT SYSTEM =====================

# ========== AI ASSISTANT CHATBOT ==========

@app.route("/api/chatbot/message", methods=["POST"])
def chatbot_message():
    """Process chatbot message and escalate if needed"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json() or {}
    user_message = data.get("message", "").strip()

    if not user_message:
        return jsonify({"error": "Message cannot be empty"}), 400

    conn = get_db()
    user_id = session["user_id"]
    session_id = data.get("session_id") or f"api_{user_id}"

    result = get_ai_response(user_message, user_id=user_id, session_id=session_id)
    ai_response = result["response"]
    should_escalate = bool(result["escalated"])

    if should_escalate:
        admin_user = conn.execute(
            "SELECT id FROM users WHERE role = 'admin' ORDER BY id LIMIT 1"
        ).fetchone()

        if admin_user:
            conn.execute("""
                INSERT INTO admin_forwarded_chats (user_id, admin_id, original_question, status)
                VALUES (?, ?, ?, 'pending')
            """, (user_id, admin_user['id'], user_message))

    conn.execute("""
        INSERT INTO chatbot_conversations (
            user_id, user_message, bot_response, escalated_to_admin, intent, confidence, session_id, metadata, toxicity_flag
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        user_id,
        user_message,
        ai_response,
        int(should_escalate),
        result["intent"],
        result["confidence"],
        session_id,
        json.dumps({
            "clarification": result["clarification"],
            "toxicity_flag": result["toxicity_flag"],
            "context": result["context"]
        }),
        int(bool(result["toxicity_flag"]))
    ))

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "ai_response": ai_response,
        "escalated": should_escalate,
        "intent": result["intent"],
        "confidence": result["confidence"],
        "clarification": result["clarification"],
        "toxicity_flag": bool(result["toxicity_flag"])
    })

# ========== ADMIN FORWARDED CHAT DASHBOARD ==========

@app.route("/admin/forwarded-chats")
def admin_forwarded_chats_list():
    """Admin view of forwarded chats from AI Assistant"""
    if not session.get("user_id") or session.get("role") != "admin":
        flash("Access denied", "danger")
        return redirect(url_for("login"))
    
    conn = get_db()
    
    # Get pending and resolved chats
    chats = conn.execute("""
        SELECT afc.id, afc.user_id, afc.original_question, afc.status, afc.created_at, u.name, u.email
        FROM admin_forwarded_chats afc
        JOIN users u ON afc.user_id = u.id
        WHERE afc.admin_id = ?
        ORDER BY afc.status ASC, afc.created_at DESC
    """, (session["user_id"],)).fetchall()
    
    conn.close()
    return render_template("admin_forwarded_chats.html", chats=chats)

@app.route("/admin/chat/<int:chat_id>")
def admin_forwarded_chat_view(chat_id):
    """Admin view of a specific forwarded chat"""
    if not session.get("user_id") or session.get("role") != "admin":
        return jsonify({"error": "Access denied"}), 403
    
    conn = get_db()
    
    chat = conn.execute("""
        SELECT afc.*, u.name, u.email
        FROM admin_forwarded_chats afc
        JOIN users u ON afc.user_id = u.id
        WHERE afc.id = ? AND afc.admin_id = ?
    """, (chat_id, session["user_id"])).fetchone()
    
    if not chat:
        conn.close()
        return jsonify({"error": "Chat not found"}), 404
    
    conn.close()
    return render_template("admin_chat_forwarded.html", chat=chat)

@app.route("/api/admin/reply-chat/<int:chat_id>", methods=["POST"])
def admin_reply_chat(chat_id):
    """Admin sends reply to forwarded chat"""
    if not session.get("user_id") or session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.get_json() or {}
    reply_message = data.get("message", "").strip()
    
    if not reply_message:
        return jsonify({"error": "Message cannot be empty"}), 400
    
    conn = get_db()
    
    # Verify admin owns this chat
    chat = conn.execute(
        "SELECT * FROM admin_forwarded_chats WHERE id = ? AND admin_id = ?",
        (chat_id, session["user_id"])
    ).fetchone()
    
    if not chat:
        conn.close()
        return jsonify({"error": "Access denied"}), 403
    
    # Store message in chat_messages table
    conn.execute("""
        INSERT INTO chat_messages (session_id, content, is_from_user, sender_id)
        VALUES (?, ?, 0, ?)
    """, (chat_id, reply_message, session["user_id"]))
    
    # Send email notification to user
    user = conn.execute("SELECT email, name FROM users WHERE id = ?", (chat['user_id'],)).fetchone()
    if user:
        try:
            send_email(
                user['email'],
                "Response from Admin",
                f"Hello {user['name']},\n\n{reply_message}\n\nBest regards,\nAdmin Team"
            )
        except Exception as e:
            print(f"Email notification error: {e}")
    
    conn.commit()
    conn.close()
    
    return jsonify({"success": True})

# ========== STUDENT-TO-STUDENT CHAT SYSTEM ==========

@app.route("/messages")
def messages():
    """Student messages page"""
    if not session.get("user_id"):
        return redirect(url_for("login"))
    
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    
    # Get accepted chat partners
    chat_partners = conn.execute("""
        SELECT u.id, u.name, u.profile_picture, 
               MAX(m.created_at) as last_message_time,
               SUM(CASE WHEN m.receiver_id = ? AND m.is_read = 0 THEN 1 ELSE 0 END) as unread_count
        FROM student_chat_messages m
        JOIN users u ON (
            (m.sender_id = ? AND m.receiver_id = u.id) OR
            (m.receiver_id = ? AND m.sender_id = u.id)
        )
        WHERE NOT EXISTS (
            SELECT 1 FROM chat_blocks 
            WHERE (blocker_id = ? AND blocked_id = u.id)
            OR (blocker_id = u.id AND blocked_id = ?)
        )
        GROUP BY u.id
        ORDER BY last_message_time DESC
    """, (session["user_id"], session["user_id"], session["user_id"], session["user_id"], session["user_id"])).fetchall()
    
    # Get pending friend requests
    pending_requests = conn.execute("""
        SELECT fr.id, u.id as user_id, u.name, u.profile_picture
        FROM friend_requests fr
        JOIN users u ON fr.sender_id = u.id
        WHERE fr.receiver_id = ? AND fr.status = 'pending'
        ORDER BY fr.created_at DESC
    """, (session["user_id"],)).fetchall()
    
    conn.close()
    return render_template("messages.html", user=user, chat_partners=chat_partners, pending_requests=pending_requests)

@app.route("/chat/<int:other_user_id>")
def chat_view(other_user_id):
    """View chat with another user"""
    if not session.get("user_id"):
        return redirect(url_for("login"))
    
    user_id = session["user_id"]
    
    if user_id == other_user_id:
        flash("Cannot chat with yourself", "warning")
        return redirect(url_for("messages"))
    
    conn = get_db()
    
    # Check if blocked
    is_blocked = conn.execute("""
        SELECT 1 FROM chat_blocks 
        WHERE (blocker_id = ? AND blocked_id = ?)
        OR (blocker_id = ? AND blocked_id = ?)
    """, (user_id, other_user_id, other_user_id, user_id)).fetchone()
    
    if is_blocked:
        flash("You cannot access this chat", "danger")
        conn.close()
        return redirect(url_for("messages"))
    
    # Get other user
    other_user = conn.execute("SELECT * FROM users WHERE id = ?", (other_user_id,)).fetchone()
    
    if not other_user:
        conn.close()
        return redirect(url_for("messages"))
    
    # Get or create friend request status
    friend_status = conn.execute("""
        SELECT * FROM friend_requests 
        WHERE (sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?)
    """, (user_id, other_user_id, other_user_id, user_id)).fetchone()
    
    # Get messages if accepted
    messages = []
    if friend_status and friend_status['status'] == 'accepted':
        messages = conn.execute("""
            SELECT * FROM student_chat_messages 
            WHERE (sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?)
            ORDER BY created_at ASC
        """, (user_id, other_user_id, other_user_id, user_id)).fetchall()
        
        # Mark messages as read
        conn.execute("""
            UPDATE student_chat_messages 
            SET is_read = 1 
            WHERE receiver_id = ? AND sender_id = ? AND is_read = 0
        """, (user_id, other_user_id))
        conn.commit()
    
    conn.close()
    
    return render_template("chat_view.html", other_user=other_user, messages=messages, friend_status=friend_status)

@app.route("/api/chat/send-request/<int:receiver_id>", methods=["POST"])
def send_friend_request(receiver_id):
    """Send a friend/chat request"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    sender_id = session["user_id"]
    
    if sender_id == receiver_id:
        return jsonify({"error": "Cannot send request to yourself"}), 400
    
    conn = get_db()
    
    # Check if request already exists
    existing = conn.execute("""
        SELECT * FROM friend_requests 
        WHERE (sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?)
    """, (sender_id, receiver_id, receiver_id, sender_id)).fetchone()
    
    if existing:
        conn.close()
        return jsonify({"error": "Request already exists"}), 400
    
    try:
        conn.execute("""
            INSERT INTO friend_requests (sender_id, receiver_id, status)
            VALUES (?, ?, 'pending')
        """, (sender_id, receiver_id))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/chat/accept-request/<int:request_id>", methods=["POST"])
def accept_friend_request(request_id):
    """Accept a friend request"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_db()
    
    friend_req = conn.execute(
        "SELECT * FROM friend_requests WHERE id = ? AND receiver_id = ?",
        (request_id, session["user_id"])
    ).fetchone()
    
    if not friend_req:
        conn.close()
        return jsonify({"error": "Request not found"}), 404
    
    conn.execute(
        "UPDATE friend_requests SET status = 'accepted', responded_at = CURRENT_TIMESTAMP WHERE id = ?",
        (request_id,)
    )
    conn.commit()
    conn.close()
    
    return jsonify({"success": True})

@app.route("/api/chat/decline-request/<int:request_id>", methods=["POST"])
def decline_friend_request(request_id):
    """Decline a friend request"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    conn = get_db()
    
    friend_req = conn.execute(
        "SELECT * FROM friend_requests WHERE id = ? AND receiver_id = ?",
        (request_id, session["user_id"])
    ).fetchone()
    
    if not friend_req:
        conn.close()
        return jsonify({"error": "Request not found"}), 404
    
    conn.execute(
        "DELETE FROM friend_requests WHERE id = ?",
        (request_id,)
    )
    conn.commit()
    conn.close()
    
    return jsonify({"success": True})

@app.route("/api/chat/send-message/<int:receiver_id>", methods=["POST"])
def send_message(receiver_id):
    """Send a message to another user"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    sender_id = session["user_id"]
    data = request.get_json() or {}
    message_text = data.get("message", "").strip()
    
    if not message_text:
        return jsonify({"error": "Message cannot be empty"}), 400
    
    conn = get_db()
    
    # Check if friend request is accepted
    friend_req = conn.execute("""
        SELECT * FROM friend_requests 
        WHERE ((sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?))
        AND status = 'accepted'
    """, (sender_id, receiver_id, receiver_id, sender_id)).fetchone()
    
    if not friend_req:
        conn.close()
        return jsonify({"error": "You must accept a friend request to chat"}), 403
    
    # Check if blocked
    is_blocked = conn.execute("""
        SELECT 1 FROM chat_blocks 
        WHERE (blocker_id = ? AND blocked_id = ?)
    """, (receiver_id, sender_id)).fetchone()
    
    if is_blocked:
        conn.close()
        return jsonify({"error": "You are blocked"}), 403
    
    # Store message
    conn.execute("""
        INSERT INTO student_chat_messages (sender_id, receiver_id, content, is_read)
        VALUES (?, ?, ?, 0)
    """, (sender_id, receiver_id, message_text))
    
    conn.commit()
    conn.close()
    
    return jsonify({"success": True})

@app.route("/api/chat/block/<int:user_id>", methods=["POST"])
def block_user(user_id):
    """Block a user"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    blocker_id = session["user_id"]
    
    if blocker_id == user_id:
        return jsonify({"error": "Cannot block yourself"}), 400
    
    conn = get_db()
    
    try:
        conn.execute("""
            INSERT OR IGNORE INTO chat_blocks (blocker_id, blocked_id)
            VALUES (?, ?)
        """, (blocker_id, user_id))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/chat/unblock/<int:user_id>", methods=["POST"])
def unblock_user(user_id):
    """Unblock a user"""
    if not session.get("user_id"):
        return jsonify({"error": "Unauthorized"}), 401
    
    blocker_id = session["user_id"]
    
    conn = get_db()
    
    conn.execute("""
        DELETE FROM chat_blocks 
        WHERE blocker_id = ? AND blocked_id = ?
    """, (blocker_id, user_id))
    
    conn.commit()
    conn.close()
    
    return jsonify({"success": True})

# ===================== RUN =====================
if __name__ == "__main__":
    socketio.run(app, debug=True, port=5000)